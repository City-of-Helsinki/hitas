from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

import pytest
from django.urls import reverse
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from rest_framework import status

from hitas.models import Apartment, ApartmentType, Building, HousingCompany, RealEstate
from hitas.tests.apis.helpers import HitasAPIClient, InvalidInput, parametrize_helper
from hitas.tests.factories import ApartmentTypeFactory, BuildingFactory, HousingCompanyFactory, RealEstateFactory


@pytest.mark.django_db
def test__api__sales_catalog(api_client: HitasAPIClient):
    housing_company: HousingCompany = HousingCompanyFactory.create()

    # Create necessary apartment types
    apartment_type_1: ApartmentType = ApartmentTypeFactory.create(value="h+k+s")
    apartment_type_2: ApartmentType = ApartmentTypeFactory.create(value="h+kt")

    url = reverse(
        "hitas:sales-catalog-validate-list",
        kwargs={
            "housing_company_uuid": housing_company.uuid.hex,
        },
    )
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    path = Path(__file__).parent.parent / "static" / "myyntihintaluettelo_esimerkki.xlsx"
    data = path.read_bytes()

    response = api_client.post(
        url,
        data=data,
        content_type=content_type,
        openapi_validate_request=False,  # cannot validate requests with bytes
    )

    assert response.status_code == status.HTTP_200_OK, response.json()
    assert response.json() == {
        "apartments": [
            {
                "acquisition_price": 275000.0,
                "apartment_number": 1,
                "apartment_type": {
                    "id": apartment_type_1.uuid.hex,
                    "value": "h+k+s",
                },
                "catalog_purchase_price": 200000.0,
                "floor": "1",
                "catalog_primary_loan_amount": 75000.0,
                "rooms": 3,
                "row": 5,
                "share_number_end": 75,
                "share_number_start": 1,
                "stair": "A",
                "surface_area": 75.0,
            },
            {
                "acquisition_price": 175000.0,
                "apartment_number": 2,
                "apartment_type": {
                    "id": apartment_type_1.uuid.hex,
                    "value": "h+k+s",
                },
                "catalog_purchase_price": 100000.0,
                "floor": "2",
                "catalog_primary_loan_amount": 75000.0,
                "rooms": 1,
                "row": 6,
                "share_number_end": 115,
                "share_number_start": 76,
                "stair": "A",
                "surface_area": 40.0,
            },
            {
                "acquisition_price": 225000.0,
                "apartment_number": 3,
                "apartment_type": {
                    "id": apartment_type_2.uuid.hex,
                    "value": "h+kt",
                },
                "catalog_purchase_price": 150000.0,
                "floor": "3",
                "catalog_primary_loan_amount": 75000.0,
                "rooms": 2,
                "row": 7,
                "share_number_end": 180,
                "share_number_start": 116,
                "stair": "A",
                "surface_area": 65.0,
            },
            {
                "acquisition_price": 350000.0,
                "apartment_number": 4,
                "apartment_type": {
                    "id": apartment_type_1.uuid.hex,
                    "value": "h+k+s",
                },
                "catalog_purchase_price": 250000.0,
                "floor": "4",
                "catalog_primary_loan_amount": 100000.0,
                "rooms": 5,
                "row": 8,
                "share_number_end": 280,
                "share_number_start": 181,
                "stair": "A",
                "surface_area": 100.0,
            },
            {
                "acquisition_price": 170000.0,
                "apartment_number": 1,
                "apartment_type": {
                    "id": apartment_type_1.uuid.hex,
                    "value": "h+k+s",
                },
                "catalog_purchase_price": 95000.0,
                "floor": "5",
                "catalog_primary_loan_amount": 75000.0,
                "rooms": 1,
                "row": 9,
                "share_number_end": 320,
                "share_number_start": 281,
                "stair": "B",
                "surface_area": 40.0,
            },
            {
                "acquisition_price": 350000.0,
                "apartment_number": 2,
                "apartment_type": {
                    "id": apartment_type_1.uuid.hex,
                    "value": "h+k+s",
                },
                "catalog_purchase_price": 250000.0,
                "floor": "6",
                "catalog_primary_loan_amount": 100000.0,
                "rooms": 5,
                "row": 10,
                "share_number_end": 420,
                "share_number_start": 321,
                "stair": "B",
                "surface_area": 100.0,
            },
            {
                "acquisition_price": 275000.0,
                "apartment_number": 3,
                "apartment_type": {
                    "id": apartment_type_1.uuid.hex,
                    "value": "h+k+s",
                },
                "catalog_purchase_price": 200000.0,
                "floor": "7",
                "catalog_primary_loan_amount": 75000.0,
                "rooms": 3,
                "row": 11,
                "share_number_end": 495,
                "share_number_start": 421,
                "stair": "B",
                "surface_area": 75.0,
            },
            {
                "acquisition_price": 235000.0,
                "apartment_number": 4,
                "apartment_type": {
                    "id": apartment_type_2.uuid.hex,
                    "value": "h+kt",
                },
                "catalog_purchase_price": 160000.0,
                "floor": "8",
                "catalog_primary_loan_amount": 75000.0,
                "rooms": 2,
                "row": 12,
                "share_number_end": 560,
                "share_number_start": 496,
                "stair": "B",
                "surface_area": 65.0,
            },
        ],
        "confirmation_date": "2021-08-06",
        "acquisition_price_limit": 2060000,
        "total_acquisition_price": 2055000,
        "total_surface_area": 560,
    }


@pytest.mark.django_db
def test__api__sales_catalog__missing_apartment_types(api_client: HitasAPIClient):
    housing_company: HousingCompany = HousingCompanyFactory.create()

    url = reverse(
        "hitas:sales-catalog-validate-list",
        kwargs={
            "housing_company_uuid": housing_company.uuid.hex,
        },
    )
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    path = Path(__file__).parent.parent / "static" / "myyntihintaluettelo_esimerkki.xlsx"
    data = path.read_bytes()

    response = api_client.post(
        url,
        data=data,
        content_type=content_type,
        openapi_validate_request=False,  # cannot validate requests with bytes
    )

    assert response.status_code == status.HTTP_400_BAD_REQUEST, response.json()
    assert response.json() == {
        "error": "bad_request",
        "fields": [
            {"field": "E5.apartment_type", "message": "Apartment type h+k+s not found"},
            {"field": "E6.apartment_type", "message": "Apartment type h+k+s not found"},
            {"field": "E7.apartment_type", "message": "Apartment type h+kt not found"},
            {"field": "E8.apartment_type", "message": "Apartment type h+k+s not found"},
            {"field": "E9.apartment_type", "message": "Apartment type h+k+s not found"},
            {"field": "E10.apartment_type", "message": "Apartment type h+k+s not found"},
            {"field": "E11.apartment_type", "message": "Apartment type h+k+s not found"},
            {"field": "E12.apartment_type", "message": "Apartment type h+kt not found"},
        ],
        "message": "Bad request",
        "reason": "Bad Request",
        "status": 400,
    }


@pytest.mark.parametrize(
    **parametrize_helper(
        {
            "Missing stair": InvalidInput(
                invalid_data={
                    "A5": None,
                },
                fields=[
                    {
                        "field": "A5.stair",
                        "message": "This field is mandatory and cannot be null.",
                    },
                ],
            ),
            "Missing floor": InvalidInput(
                invalid_data={
                    "B5": None,
                },
                fields=[
                    {
                        "field": "B5.floor",
                        "message": "This field is mandatory and cannot be null.",
                    },
                ],
            ),
            "Missing apartment_number": InvalidInput(
                invalid_data={
                    "C5": None,
                },
                fields=[
                    {
                        "field": "C5.apartment_number",
                        "message": "This field is mandatory and cannot be null.",
                    },
                ],
            ),
            "Missing rooms": InvalidInput(
                invalid_data={
                    "D5": None,
                },
                fields=[
                    {
                        "field": "D5.rooms",
                        "message": "This field is mandatory and cannot be null.",
                    },
                ],
            ),
            "Missing apartment_type": InvalidInput(
                invalid_data={
                    "E5": None,
                },
                fields=[
                    {
                        "field": "E5.apartment_type",
                        "message": "This field is mandatory and cannot be null.",
                    },
                ],
            ),
            "Missing surface_area": InvalidInput(
                invalid_data={
                    "F5": None,
                },
                fields=[
                    {
                        "field": "F5.surface_area",
                        "message": "This field is mandatory and cannot be null.",
                    },
                ],
            ),
            "Missing share_number_start": InvalidInput(
                invalid_data={
                    "G5": None,
                },
                fields=[
                    {
                        "field": "G5.share_number_start",
                        "message": "This field is mandatory and cannot be null.",
                    },
                ],
            ),
            "Missing share_number_end": InvalidInput(
                invalid_data={
                    "H5": None,
                },
                fields=[
                    {
                        "field": "H5.share_number_end",
                        "message": "This field is mandatory and cannot be null.",
                    },
                ],
            ),
            "Missing catalog_purchase_price": InvalidInput(
                invalid_data={
                    "I5": None,
                },
                fields=[
                    {
                        "field": "I5.catalog_purchase_price",
                        "message": "This field is mandatory and cannot be null.",
                    },
                ],
            ),
            "Missing catalog_primary_loan_amount": InvalidInput(
                invalid_data={
                    "J5": None,
                },
                fields=[
                    {
                        "field": "J5.catalog_primary_loan_amount",
                        "message": "This field is mandatory and cannot be null.",
                    },
                ],
            ),
            "Missing acquisition_price": InvalidInput(
                invalid_data={
                    "K5": None,
                },
                fields=[
                    {
                        "field": "K5.acquisition_price",
                        "message": "This field is mandatory and cannot be null.",
                    },
                ],
            ),
            "Missing confirmation_date": InvalidInput(
                invalid_data={
                    "F2": None,
                },
                fields=[
                    {
                        "field": "F2.confirmation_date",
                        "message": "This field is mandatory and cannot be null.",
                    },
                ],
            ),
            "Missing acquisition_price_limit": InvalidInput(
                invalid_data={
                    "E3": None,
                },
                fields=[
                    {
                        "field": "E3.acquisition_price_limit",
                        "message": "This field is mandatory and cannot be null.",
                    },
                ],
            ),
            "Missing total_surface_area": InvalidInput(
                invalid_data={
                    "F13": None,
                },
                fields=[
                    {
                        "field": "F13.total_surface_area",
                        "message": "This field is mandatory and cannot be null.",
                    },
                ],
            ),
            "Missing total_acquisition_price": InvalidInput(
                invalid_data={
                    "K13": None,
                },
                fields=[
                    {
                        "field": "K13.total_acquisition_price",
                        "message": "This field is mandatory and cannot be null.",
                    },
                ],
            ),
            "Share overlap single start": InvalidInput(
                invalid_data={
                    "G6": 75,
                },
                fields=[
                    {
                        "field": "H5.share_number_end",
                        "message": "Overlapping shares with apartment on row 6: 75",
                    },
                    {
                        "field": "G6.share_number_start",
                        "message": "Overlapping shares with apartment on row 5: 75",
                    },
                ],
            ),
            "Share overlap single end": InvalidInput(
                invalid_data={
                    "H5": 76,
                },
                fields=[
                    {
                        "field": "H5.share_number_end",
                        "message": "Overlapping shares with apartment on row 6: 76",
                    },
                    {
                        "field": "G6.share_number_start",
                        "message": "Overlapping shares with apartment on row 5: 76",
                    },
                ],
            ),
            "Share overlap multiple": InvalidInput(
                invalid_data={
                    "G6": 70,
                },
                fields=[
                    {
                        "field": "G5.share_number_start",
                        "message": "Overlapping shares with apartment on row 6: 70-75",
                    },
                    {
                        "field": "G6.share_number_start",
                        "message": "Overlapping shares with apartment on row 5: 70-75",
                    },
                    {
                        "field": "H5.share_number_end",
                        "message": "Overlapping shares with apartment on row 6: 70-75",
                    },
                    {
                        "field": "H6.share_number_end",
                        "message": "Overlapping shares with apartment on row 5: 70-75",
                    },
                ],
            ),
            "total_acquisition_price over acquisition_price_limit": InvalidInput(
                invalid_data={
                    "E3": 1,  # acquisition_price_limit
                    "K13": 2,  # total_acquisition_price
                },
                fields=[
                    {
                        "field": "K13.total_acquisition_price",
                        "message": "'total_acquisition_price' is higher than 'acquisition_price_limit'.",
                    },
                ],
            ),
        },
    )
)
@pytest.mark.django_db
def test__api__sales_catalog__invalid_data(api_client: HitasAPIClient, invalid_data, fields):
    housing_company: HousingCompany = HousingCompanyFactory.create()

    # Create necessary apartment types
    ApartmentTypeFactory.create(value="h+k+s")
    ApartmentTypeFactory.create(value="h+kt")

    url = reverse(
        "hitas:sales-catalog-validate-list",
        kwargs={
            "housing_company_uuid": housing_company.uuid.hex,
        },
    )
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    path = Path(__file__).parent.parent / "static" / "myyntihintaluettelo_esimerkki.xlsx"
    wb: Workbook = load_workbook(path, data_only=True)
    ws: Worksheet = wb.worksheets[0]

    for key, value in invalid_data.items():
        ws[key] = value

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        data: bytes = tmp.read()

    response = api_client.post(
        url,
        data=data,
        content_type=content_type,
        openapi_validate_request=False,  # cannot validate requests with bytes
    )

    assert response.status_code == status.HTTP_400_BAD_REQUEST, response.json()
    assert response.json() == {
        "error": "bad_request",
        "fields": fields,
        "message": "Bad request",
        "reason": "Bad Request",
        "status": 400,
    }


@pytest.mark.django_db
def test__api__sales_catalog__create(api_client: HitasAPIClient):
    housing_company: HousingCompany = HousingCompanyFactory.create()

    # Create necessary apartment types
    apartment_type_1: ApartmentType = ApartmentTypeFactory.create(value="h+k+s")
    apartment_type_2: ApartmentType = ApartmentTypeFactory.create(value="h+kt")

    url = reverse(
        "hitas:sales-catalog-create-list",
        kwargs={
            "housing_company_uuid": housing_company.uuid.hex,
        },
    )

    data = [
        {
            "apartment_number": 1,
            "apartment_type": apartment_type_1.uuid.hex,
            "catalog_purchase_price": 200000.0,
            "floor": "1",
            "catalog_primary_loan_amount": 75000.0,
            "rooms": 3,
            "share_number_end": 75,
            "share_number_start": 1,
            "stair": "A",
            "surface_area": 75.0,
        },
        {
            "apartment_number": 2,
            "apartment_type": apartment_type_1.uuid.hex,
            "catalog_purchase_price": 100000.0,
            "floor": "2",
            "catalog_primary_loan_amount": 75000.0,
            "rooms": 1,
            "share_number_end": 115,
            "share_number_start": 76,
            "stair": "A",
            "surface_area": 40.0,
        },
        {
            "apartment_number": 3,
            "apartment_type": apartment_type_2.uuid.hex,
            "catalog_purchase_price": 150000.0,
            "floor": "3",
            "catalog_primary_loan_amount": 75000.0,
            "rooms": 2,
            "share_number_end": 180,
            "share_number_start": 116,
            "stair": "A",
            "surface_area": 65.0,
        },
        {
            "apartment_number": 4,
            "apartment_type": apartment_type_1.uuid.hex,
            "catalog_purchase_price": 250000.0,
            "floor": "4",
            "catalog_primary_loan_amount": 100000.0,
            "rooms": 5,
            "share_number_end": 280,
            "share_number_start": 181,
            "stair": "A",
            "surface_area": 100.0,
        },
        {
            "apartment_number": 1,
            "apartment_type": apartment_type_1.uuid.hex,
            "catalog_purchase_price": 95000.0,
            "floor": "5",
            "catalog_primary_loan_amount": 75000.0,
            "rooms": 1,
            "share_number_end": 320,
            "share_number_start": 281,
            "stair": "B",
            "surface_area": 40.0,
        },
        {
            "apartment_number": 2,
            "apartment_type": apartment_type_1.uuid.hex,
            "catalog_purchase_price": 250000.0,
            "floor": "6",
            "catalog_primary_loan_amount": 100000.0,
            "rooms": 5,
            "share_number_end": 420,
            "share_number_start": 321,
            "stair": "B",
            "surface_area": 100.0,
        },
        {
            "apartment_number": 3,
            "apartment_type": apartment_type_1.uuid.hex,
            "catalog_purchase_price": 200000.0,
            "floor": "7",
            "catalog_primary_loan_amount": 75000.0,
            "rooms": 3,
            "share_number_end": 495,
            "share_number_start": 421,
            "stair": "B",
            "surface_area": 75.0,
        },
        {
            "apartment_number": 4,
            "apartment_type": apartment_type_2.uuid.hex,
            "catalog_purchase_price": 160000.0,
            "floor": "8",
            "catalog_primary_loan_amount": 75000.0,
            "rooms": 2,
            "share_number_end": 560,
            "share_number_start": 496,
            "stair": "B",
            "surface_area": 65.0,
        },
    ]

    response = api_client.post(url, data=data, format="json")

    assert response.status_code == status.HTTP_201_CREATED, response.json()

    apartments: list[Apartment] = list(Apartment.objects.all())
    assert len(apartments) == 8
    assert apartments[0].stair == "A"
    assert apartments[0].apartment_number == 1
    assert apartments[1].stair == "A"
    assert apartments[1].apartment_number == 2
    assert apartments[2].stair == "A"
    assert apartments[2].apartment_number == 3
    assert apartments[3].stair == "A"
    assert apartments[3].apartment_number == 4
    assert apartments[4].stair == "B"
    assert apartments[4].apartment_number == 1
    assert apartments[5].stair == "B"
    assert apartments[5].apartment_number == 2
    assert apartments[6].stair == "B"
    assert apartments[6].apartment_number == 3
    assert apartments[7].stair == "B"
    assert apartments[7].apartment_number == 4

    # Real estate and building created automatically
    real_estates: list[RealEstate] = list(housing_company.real_estates.all())
    assert len(real_estates) == 1
    buildings: list[Building] = list(real_estates[0].buildings.all())
    assert len(buildings) == 1


@pytest.mark.django_db
def test__api__sales_catalog__create__real_estate_and_building_exist(api_client: HitasAPIClient):
    housing_company: HousingCompany = HousingCompanyFactory.create()
    real_estate: RealEstate = RealEstateFactory.create(housing_company=housing_company)
    building: Building = BuildingFactory.create(real_estate=real_estate)

    # Create necessary apartment types
    apartment_type: ApartmentType = ApartmentTypeFactory.create(value="h+k+s")

    url = reverse(
        "hitas:sales-catalog-create-list",
        kwargs={
            "housing_company_uuid": housing_company.uuid.hex,
        },
    )

    data = [
        {
            "apartment_number": 1,
            "apartment_type": apartment_type.uuid.hex,
            "catalog_purchase_price": 200000.0,
            "floor": "1",
            "catalog_primary_loan_amount": 75000.0,
            "rooms": 3,
            "share_number_end": 75,
            "share_number_start": 1,
            "stair": "A",
            "surface_area": 75.0,
        },
    ]

    response = api_client.post(url, data=data, format="json")

    assert response.status_code == status.HTTP_201_CREATED, response.json()

    apartments: list[Apartment] = list(Apartment.objects.all())
    assert len(apartments) == 1

    # Use existing real estate and building
    real_estates: list[RealEstate] = list(housing_company.real_estates.all())
    assert len(real_estates) == 1
    assert real_estates[0] == real_estate
    buildings: list[Building] = list(real_estates[0].buildings.all())
    assert len(buildings) == 1
    assert buildings[0] == building


@pytest.mark.parametrize(
    **parametrize_helper(
        {
            "apartment_number cannot be null": InvalidInput(
                invalid_data={
                    "apartment_number": None,
                },
                fields=[
                    {
                        "field": "0.apartment_number",
                        "message": "This field is mandatory and cannot be null.",
                    },
                ],
            ),
            "stair cannot be null": InvalidInput(
                invalid_data={
                    "stair": None,
                },
                fields=[
                    {
                        "field": "0.stair",
                        "message": "This field is mandatory and cannot be null.",
                    },
                ],
            ),
        },
    )
)
@pytest.mark.django_db
def test__api__sales_catalog__create__invalid_data(api_client: HitasAPIClient, invalid_data, fields):
    housing_company: HousingCompany = HousingCompanyFactory.create()

    # Create necessary apartment types
    apartment_type: ApartmentType = ApartmentTypeFactory.create(value="h+k+s")

    url = reverse(
        "hitas:sales-catalog-create-list",
        kwargs={
            "housing_company_uuid": housing_company.uuid.hex,
        },
    )

    data = {
        "apartment_number": 1,
        "apartment_type": apartment_type.uuid.hex,
        "catalog_purchase_price": 200000.0,
        "floor": "1",
        "catalog_primary_loan_amount": 75000.0,
        "rooms": 3,
        "share_number_end": 75,
        "share_number_start": 1,
        "stair": "A",
        "surface_area": 75.0,
    }
    data.update(invalid_data)

    response = api_client.post(url, data=[data], format="json", openapi_validate_request=False)

    assert response.status_code == status.HTTP_400_BAD_REQUEST, response.json()
    assert response.json() == {
        "error": "bad_request",
        "fields": fields,
        "message": "Bad request",
        "reason": "Bad Request",
        "status": 400,
    }

    # Real estate is not created
    real_estates: list[RealEstate] = list(housing_company.real_estates.all())
    assert len(real_estates) == 0


@pytest.mark.django_db
def test__api__sales_catalog__create__invalid_data__some_in_list(api_client: HitasAPIClient):
    housing_company: HousingCompany = HousingCompanyFactory.create()

    # Create necessary apartment types
    apartment_type: ApartmentType = ApartmentTypeFactory.create(value="h+k+s")

    url = reverse(
        "hitas:sales-catalog-create-list",
        kwargs={
            "housing_company_uuid": housing_company.uuid.hex,
        },
    )

    invalid_data = {
        "apartment_number": None,
    }

    data_1: dict[str, Any] = {
        "apartment_number": 1,
        "apartment_type": apartment_type.uuid.hex,
        "catalog_purchase_price": 200000.0,
        "floor": "1",
        "catalog_primary_loan_amount": 75000.0,
        "rooms": 3,
        "share_number_end": 75,
        "share_number_start": 1,
        "stair": "A",
        "surface_area": 75.0,
    }
    data_2 = data_1.copy()
    data_3 = data_1.copy()
    data_1.update(invalid_data)
    data_3.update(invalid_data)

    response = api_client.post(url, data=[data_1, data_2, data_3], format="json", openapi_validate_request=False)

    assert response.status_code == status.HTTP_400_BAD_REQUEST, response.json()
    assert response.json() == {
        "error": "bad_request",
        "fields": [
            {
                "field": "0.apartment_number",
                "message": "This field is mandatory and cannot be null.",
            },
            {
                "field": "2.apartment_number",
                "message": "This field is mandatory and cannot be null.",
            },
        ],
        "message": "Bad request",
        "reason": "Bad Request",
        "status": 400,
    }

    # Real estate is not created
    real_estates: list[RealEstate] = list(housing_company.real_estates.all())
    assert len(real_estates) == 0


@pytest.mark.django_db
def test__api__sales_catalog__create__empty(api_client: HitasAPIClient):
    housing_company: HousingCompany = HousingCompanyFactory.create()

    url = reverse(
        "hitas:sales-catalog-create-list",
        kwargs={
            "housing_company_uuid": housing_company.uuid.hex,
        },
    )

    response = api_client.post(url, data=[], format="json", openapi_validate_request=False)

    assert response.status_code == status.HTTP_400_BAD_REQUEST, response.json()
    assert response.json() == {
        "error": "bad_request",
        "fields": [
            {
                "field": "non_field_errors",
                "message": "This list may not be empty.",
            },
        ],
        "message": "Bad request",
        "reason": "Bad Request",
        "status": 400,
    }

    # Real estate is not created
    real_estates: list[RealEstate] = list(housing_company.real_estates.all())
    assert len(real_estates) == 0
