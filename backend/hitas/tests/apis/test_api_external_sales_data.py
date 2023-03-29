from pathlib import Path
from tempfile import NamedTemporaryFile

import pytest
from django.urls import reverse
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from rest_framework import status

from hitas.models import ExternalSalesData
from hitas.models.external_sales_data import CostAreaData, ExternalSalesDataType, QuarterData
from hitas.tests.apis.helpers import HitasAPIClient, InvalidInput, parametrize_helper
from hitas.tests.factories import HitasPostalCodeFactory, HousingCompanyFactory

# Create tests


@pytest.mark.django_db
def test__api__external_sales_data__create(api_client: HitasAPIClient):
    url = reverse("hitas:external-sales-data-list")
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    postal_codes = [
        "00180",
        "00220",
        "00280",
        "00300",
        "00310",
        "00540",
        "00570",
        "00580",
        "00590",
        "00650",
        "00680",
        "00690",
        "00850",
        "00870",
    ]

    for postal_code in postal_codes:
        HousingCompanyFactory.create(postal_code__value=postal_code)

    path = Path(__file__).parent.parent / "static" / "tilastokeskuksen_esimerkki.xlsx"
    data = path.read_bytes()

    response = api_client.post(
        url,
        data=data,
        content_type=content_type,
        openapi_validate_request=False,  # cannot validate requests with bytes
    )

    assert response.status_code == status.HTTP_201_CREATED, response.json()
    assert response.json() == {
        "calculation_quarter": "2022Q4",
        "quarter_1": {
            "quarter": "2022Q1",
            "areas": [
                {"postal_code": "00180", "sale_count": 77, "price": 8304},
                {"postal_code": "00220", "sale_count": 19, "price": 8255},
                {"postal_code": "00280", "sale_count": 19, "price": 6747},
                {"postal_code": "00300", "sale_count": 17, "price": 5577},
                {"postal_code": "00310", "sale_count": 8, "price": 4222},
                {"postal_code": "00540", "sale_count": 18, "price": 7684},
                {"postal_code": "00570", "sale_count": 34, "price": 6527},
                {"postal_code": "00580", "sale_count": 29, "price": 8156},
                {"postal_code": "00650", "sale_count": 21, "price": 3810},
                {"postal_code": "00680", "sale_count": 12, "price": 4072},
                {"postal_code": "00690", "sale_count": 10, "price": 3608},
                {"postal_code": "00870", "sale_count": 20, "price": 3556},
            ],
        },
        "quarter_2": {
            "quarter": "2022Q2",
            "areas": [
                {"postal_code": "00180", "sale_count": 82, "price": 8101},
                {"postal_code": "00220", "sale_count": 22, "price": 7369},
                {"postal_code": "00280", "sale_count": 16, "price": 6573},
                {"postal_code": "00300", "sale_count": 26, "price": 5745},
                {"postal_code": "00310", "sale_count": 7, "price": 4674},
                {"postal_code": "00540", "sale_count": 26, "price": 8841},
                {"postal_code": "00570", "sale_count": 31, "price": 6642},
                {"postal_code": "00580", "sale_count": 24, "price": 8399},
                {"postal_code": "00590", "sale_count": 12, "price": 5073},
                {"postal_code": "00650", "sale_count": 22, "price": 4108},
                {"postal_code": "00680", "sale_count": 10, "price": 3841},
                {"postal_code": "00690", "sale_count": 10, "price": 3931},
                {"postal_code": "00870", "sale_count": 18, "price": 3316},
            ],
        },
        "quarter_3": {
            "quarter": "2022Q3",
            "areas": [
                {"postal_code": "00180", "sale_count": 62, "price": 8504},
                {"postal_code": "00220", "sale_count": 22, "price": 8842},
                {"postal_code": "00280", "sale_count": 16, "price": 6341},
                {"postal_code": "00300", "sale_count": 15, "price": 5807},
                {"postal_code": "00540", "sale_count": 13, "price": 8941},
                {"postal_code": "00570", "sale_count": 25, "price": 6589},
                {"postal_code": "00580", "sale_count": 21, "price": 7808},
                {"postal_code": "00650", "sale_count": 18, "price": 5129},
                {"postal_code": "00690", "sale_count": 6, "price": 3827},
                {"postal_code": "00850", "sale_count": 11, "price": 4566},
                {"postal_code": "00870", "sale_count": 19, "price": 3521},
            ],
        },
        "quarter_4": {
            "quarter": "2022Q4",
            "areas": [
                {"postal_code": "00180", "sale_count": 33, "price": 7919},
                {"postal_code": "00220", "sale_count": 17, "price": 8115},
                {"postal_code": "00570", "sale_count": 13, "price": 6129},
                {"postal_code": "00580", "sale_count": 13, "price": 7777},
            ],
        },
    }

    sales_data = list(ExternalSalesData.objects.all())
    assert len(sales_data) == 1
    assert sales_data[0].calculation_quarter == "2022Q4"


@pytest.mark.django_db
def test__api__external_sales_data__create__missing_postal_codes(api_client: HitasAPIClient):
    url = reverse("hitas:external-sales-data-list")
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    path = Path(__file__).parent.parent / "static" / "tilastokeskuksen_esimerkki.xlsx"
    data = path.read_bytes()

    response = api_client.post(
        url,
        data=data,
        content_type=content_type,
        openapi_validate_request=False,  # cannot validate requests with bytes
    )

    assert response.status_code == status.HTTP_201_CREATED, response.json()
    assert response.json() == {
        "calculation_quarter": "2022Q4",
        "quarter_1": {
            "quarter": "2022Q1",
            "areas": [],
        },
        "quarter_2": {
            "quarter": "2022Q2",
            "areas": [],
        },
        "quarter_3": {
            "quarter": "2022Q3",
            "areas": [],
        },
        "quarter_4": {
            "quarter": "2022Q4",
            "areas": [],
        },
    }

    sales_data = list(ExternalSalesData.objects.all())
    assert len(sales_data) == 1
    assert sales_data[0].calculation_quarter == "2022Q4"


@pytest.mark.django_db
def test__api__external_sales_data__create__postal_codes_not_on_housing_companies(api_client: HitasAPIClient):
    url = reverse("hitas:external-sales-data-list")
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    postal_codes = [
        "00180",
        "00220",
        "00280",
        "00300",
        "00310",
        "00540",
        "00570",
        "00580",
        "00590",
        "00650",
        "00680",
        "00690",
        "00850",
        "00870",
    ]

    for postal_code in postal_codes:
        HitasPostalCodeFactory.create(value=postal_code)

    path = Path(__file__).parent.parent / "static" / "tilastokeskuksen_esimerkki.xlsx"
    data = path.read_bytes()

    response = api_client.post(
        url,
        data=data,
        content_type=content_type,
        openapi_validate_request=False,  # cannot validate requests with bytes
    )

    assert response.status_code == status.HTTP_201_CREATED, response.json()
    assert response.json() == {
        "calculation_quarter": "2022Q4",
        "quarter_1": {
            "quarter": "2022Q1",
            "areas": [],
        },
        "quarter_2": {
            "quarter": "2022Q2",
            "areas": [],
        },
        "quarter_3": {
            "quarter": "2022Q3",
            "areas": [],
        },
        "quarter_4": {
            "quarter": "2022Q4",
            "areas": [],
        },
    }

    sales_data = list(ExternalSalesData.objects.all())
    assert len(sales_data) == 1
    assert sales_data[0].calculation_quarter == "2022Q4"


@pytest.mark.parametrize(
    **parametrize_helper(
        {
            "Missing postal code": InvalidInput(
                invalid_data={
                    "A5": None,
                },
                fields=[
                    {
                        "field": "A5.postal_code",
                        "message": "This field is mandatory and cannot be null.",
                    },
                ],
            ),
            "Invalid postal code": InvalidInput(
                invalid_data={
                    "A5": "1234 Ruoholahti",
                },
                fields=[
                    {
                        "field": "A5.postal_code",
                        "message": "'1234' is not a valid postal code.",
                    },
                ],
            ),
            "Missing quarter_1": InvalidInput(
                invalid_data={
                    "H3": None,
                },
                fields=[
                    {
                        "field": "H3.quarter_1",
                        "message": "This field is mandatory and cannot be null.",
                    },
                ],
            ),
            "Invalid quarter_1": InvalidInput(
                invalid_data={
                    "H3": "foobar",
                },
                fields=[
                    {
                        "field": "H3.quarter_1",
                        "message": "'FOOBAR' is not a valid quarter.",
                    },
                ],
            ),
            "Missing quarter_2": InvalidInput(
                invalid_data={
                    "J3": None,
                },
                fields=[
                    {
                        "field": "J3.quarter_2",
                        "message": "This field is mandatory and cannot be null.",
                    },
                ],
            ),
            "Invalid quarter_2": InvalidInput(
                invalid_data={
                    "J3": "foobar",
                },
                fields=[
                    {
                        "field": "J3.quarter_2",
                        "message": "'FOOBAR' is not a valid quarter.",
                    },
                ],
            ),
            "Missing quarter_3": InvalidInput(
                invalid_data={
                    "L3": None,
                },
                fields=[
                    {
                        "field": "L3.quarter_3",
                        "message": "This field is mandatory and cannot be null.",
                    },
                ],
            ),
            "Invalid quarter_3": InvalidInput(
                invalid_data={
                    "L3": "foobar",
                },
                fields=[
                    {
                        "field": "L3.quarter_3",
                        "message": "'FOOBAR' is not a valid quarter.",
                    },
                ],
            ),
            "Missing quarter_4": InvalidInput(
                invalid_data={
                    "N3": None,
                },
                fields=[
                    {
                        "field": "N3.quarter_4",
                        "message": "This field is mandatory and cannot be null.",
                    },
                ],
            ),
            "Invalid quarter_4": InvalidInput(
                invalid_data={
                    "N3": "foobar",
                },
                fields=[
                    {
                        "field": "N3.quarter_4",
                        "message": "'FOOBAR' is not a valid quarter.",
                    },
                ],
            ),
            "Invalid quarter, no Q5": InvalidInput(
                invalid_data={
                    "H3": "2022Q5",
                },
                fields=[
                    {
                        "field": "H3.quarter_1",
                        "message": "'2022Q5' is not a valid quarter.",
                    },
                ],
            ),
            "Invalid quarter, improper quarter sign": InvalidInput(
                invalid_data={
                    "H3": "2022W1",
                },
                fields=[
                    {
                        "field": "H3.quarter_1",
                        "message": "'2022W1' is not a valid quarter.",
                    },
                ],
            ),
            "Invalid quarter, too short": InvalidInput(
                invalid_data={
                    "H3": "2022Q",
                },
                fields=[
                    {
                        "field": "H3.quarter_1",
                        "message": "Ensure this field has at least 6 characters.",
                    },
                ],
            ),
            "Invalid price": InvalidInput(
                invalid_data={
                    "H5": "foo",
                },
                fields=[
                    {
                        "field": "H5.quarter_1_price",
                        "message": "A valid integer is required.",
                    },
                ],
            ),
            "Invalid sale_count": InvalidInput(
                invalid_data={
                    "I5": "foo",
                },
                fields=[
                    {
                        "field": "I5.quarter_1_sale_count",
                        "message": "A valid integer is required.",
                    },
                ],
            ),
        },
    )
)
@pytest.mark.django_db
def test__api__external_sales_data__create__invalid_data(api_client: HitasAPIClient, invalid_data, fields):
    url = reverse("hitas:external-sales-data-list")
    content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    postal_codes = [
        "00180",
        "00220",
        "00280",
        "00300",
        "00310",
        "00540",
        "00570",
        "00580",
        "00590",
        "00650",
        "00680",
        "00690",
        "00850",
        "00870",
    ]

    for postal_code in postal_codes:
        HousingCompanyFactory.create(postal_code__value=postal_code)

    path = Path(__file__).parent.parent / "static" / "tilastokeskuksen_esimerkki.xlsx"
    wb: Workbook = load_workbook(path, data_only=True)
    ws: Worksheet = wb.worksheets[0]

    for key, value in invalid_data.items():
        ws[key] = value

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        data: bytes = tmp.read()

    response = api_client.post(
        url,
        data=data,
        content_type=content_type,
        openapi_validate_request=False,  # cannot validate requests with bytes
    )

    assert response.status_code == status.HTTP_400_BAD_REQUEST, response.json()
    assert response.json() == {
        "error": "bad_request",
        "fields": fields,
        "message": "Bad request",
        "reason": "Bad Request",
        "status": 400,
    }


# Retrieve tests


@pytest.mark.django_db
def test__api__external_sales_data__retrieve(api_client: HitasAPIClient):
    data = ExternalSalesDataType(
        quarter_1=QuarterData(quarter="2022Q1", areas=[CostAreaData(postal_code="00000", sale_count=1, price=2)]),
        quarter_2=QuarterData(quarter="2022Q2", areas=[CostAreaData(postal_code="00000", sale_count=3, price=4)]),
        quarter_3=QuarterData(quarter="2022Q3", areas=[CostAreaData(postal_code="00000", sale_count=5, price=6)]),
        quarter_4=QuarterData(quarter="2022Q4", areas=[CostAreaData(postal_code="00000", sale_count=7, price=8)]),
    )
    ExternalSalesData.objects.create(calculation_quarter="2022Q4", **data)

    url = reverse("hitas:external-sales-data-detail", args=["2022Q4"])
    response = api_client.get(url, format="json")

    assert response.status_code == status.HTTP_200_OK, response.json()
    assert response.json() == {**{"calculation_quarter": "2022Q4"}, **data}


@pytest.mark.django_db
def test__api__external_sales_data__retrieve__not_found(api_client: HitasAPIClient):
    url = reverse("hitas:external-sales-data-detail", args=["2022Q4"])
    response = api_client.get(url, format="json")

    assert response.status_code == status.HTTP_404_NOT_FOUND, response.json()
    assert response.json() == {
        "error": "external_sales_data_not_found",
        "message": "External sales data not found",
        "reason": "Not Found",
        "status": 404,
    }
