import datetime
from io import BytesIO
from urllib.parse import urlencode

import pytest
from django.http import HttpResponse
from django.urls import reverse
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from rest_framework import status

from hitas.models import (
    Apartment,
    ApartmentSale,
    HousingCompany,
    Owner,
    Ownership,
    ThirtyYearRegulationResults,
    ThirtyYearRegulationResultsRow,
)
from hitas.models.housing_company import HitasType, RegulationStatus
from hitas.models.thirty_year_regulation import FullSalesData, RegulationResult
from hitas.tests.apis.helpers import HitasAPIClient
from hitas.tests.factories import (
    ApartmentFactory,
    ApartmentSaleFactory,
    HousingCompanyFactory,
    OwnerFactory,
    OwnershipFactory,
)

# Sales report


@pytest.mark.django_db
def test__api__sales_report__single_sale(api_client: HitasAPIClient):
    # Create sales in the report interval
    sale: ApartmentSale = ApartmentSaleFactory.create(
        purchase_date=datetime.date(2020, 1, 1),
        purchase_price=50_000,
        apartment_share_of_housing_company_loans=10_000,
        apartment__surface_area=100,
        apartment__building__real_estate__housing_company__postal_code__value="00001",
        apartment__building__real_estate__housing_company__postal_code__cost_area=1,
    )

    data = {
        "start_date": "2020-01-01",
        "end_date": "2020-01-31",
    }
    url = reverse("hitas:sales-report-list") + "?" + urlencode(data)
    response: HttpResponse = api_client.get(url)

    assert response.status_code == status.HTTP_200_OK

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert list(worksheet.values) == [
        (
            "Kalleusalue",
            "Postinumero",
            "Osoite",
            "Ilmoituspäivä",
            "Kauppapäivä",
            "Kaupan neliöhinta",
            "Velaton neliöhinta",
        ),
        (
            sale.apartment.postal_code.cost_area,
            sale.apartment.postal_code.value,
            sale.apartment.address,
            datetime.datetime.fromisoformat(sale.notification_date.isoformat()),
            datetime.datetime.fromisoformat(sale.purchase_date.isoformat()),
            500,  # 50_000 / 100
            600,  # (50_000 + 10_000) / 100
        ),
        (None, None, None, None, None, None, None),
        (None, None, None, "Kaikki kaupat", "Lukumäärä", 1, 1),
        (None, None, None, None, "Keskiarvo", 500, 600),
        (None, None, None, None, "Maksimi", 500, 600),
        (None, None, None, None, "Minimi", 500, 600),
        (None, None, None, None, None, None, None),
        (None, None, None, "Kalleusalue 1", "Lukumäärä", 1, 1),
        (None, None, None, None, "Keskiarvo", 500, 600),
        (None, None, None, None, "Maksimi", 500, 600),
        (None, None, None, None, "Minimi", 500, 600),
        (None, None, None, None, None, None, None),
        (None, None, None, "Kalleusalue 2", "Lukumäärä", 0, 0),
        (None, None, None, None, "Keskiarvo", 0, 0),
        (None, None, None, None, "Maksimi", 0, 0),
        (None, None, None, None, "Minimi", 0, 0),
        (None, None, None, None, None, None, None),
        (None, None, None, "Kalleusalue 3", "Lukumäärä", 0, 0),
        (None, None, None, None, "Keskiarvo", 0, 0),
        (None, None, None, None, "Maksimi", 0, 0),
        (None, None, None, None, "Minimi", 0, 0),
        (None, None, None, None, None, None, None),
        (None, None, None, "Kalleusalue 4", "Lukumäärä", 0, 0),
        (None, None, None, None, "Keskiarvo", 0, 0),
        (None, None, None, None, "Maksimi", 0, 0),
        (None, None, None, None, "Minimi", 0, 0),
    ]


@pytest.mark.django_db
def test__api__sales_report__multiple_sales(api_client: HitasAPIClient):
    # Create sales in the report interval
    sale_1: ApartmentSale = ApartmentSaleFactory.create(
        purchase_date=datetime.date(2020, 1, 1),
        purchase_price=50_000,
        apartment_share_of_housing_company_loans=10_000,
        apartment__surface_area=100,
        apartment__building__real_estate__housing_company__postal_code__value="00001",
        apartment__building__real_estate__housing_company__postal_code__cost_area=1,
    )
    sale_2: ApartmentSale = ApartmentSaleFactory.create(
        purchase_date=datetime.date(2020, 1, 15),
        purchase_price=130_000,
        apartment_share_of_housing_company_loans=100,
        apartment__surface_area=100,
        apartment__building__real_estate__housing_company__postal_code__value="00002",
        apartment__building__real_estate__housing_company__postal_code__cost_area=1,
    )

    data = {
        "start_date": "2020-01-01",
        "end_date": "2020-01-31",
    }
    url = reverse("hitas:sales-report-list") + "?" + urlencode(data)
    response: HttpResponse = api_client.get(url)

    assert response.status_code == status.HTTP_200_OK

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert list(worksheet.values) == [
        (
            "Kalleusalue",
            "Postinumero",
            "Osoite",
            "Ilmoituspäivä",
            "Kauppapäivä",
            "Kaupan neliöhinta",
            "Velaton neliöhinta",
        ),
        (
            sale_1.apartment.postal_code.cost_area,
            sale_1.apartment.postal_code.value,
            sale_1.apartment.address,
            datetime.datetime.fromisoformat(sale_1.notification_date.isoformat()),
            datetime.datetime.fromisoformat(sale_1.purchase_date.isoformat()),
            500,  # 50_000 / 100
            600,  # (50_000 + 10_000) / 100
        ),
        (
            sale_2.apartment.postal_code.cost_area,
            sale_2.apartment.postal_code.value,
            sale_2.apartment.address,
            datetime.datetime.fromisoformat(sale_2.notification_date.isoformat()),
            datetime.datetime.fromisoformat(sale_2.purchase_date.isoformat()),
            1300,  # 130_000 / 100
            1301,  # (130_000 + 100) / 100
        ),
        (None, None, None, None, None, None, None),
        (None, None, None, "Kaikki kaupat", "Lukumäärä", 2, 2),
        (None, None, None, None, "Keskiarvo", 900, 950.5),
        (None, None, None, None, "Maksimi", 1300, 1301),
        (None, None, None, None, "Minimi", 500, 600),
        (None, None, None, None, None, None, None),
        (None, None, None, "Kalleusalue 1", "Lukumäärä", 2, 2),
        (None, None, None, None, "Keskiarvo", 900, 950.5),
        (None, None, None, None, "Maksimi", 1300, 1301),
        (None, None, None, None, "Minimi", 500, 600),
        (None, None, None, None, None, None, None),
        (None, None, None, "Kalleusalue 2", "Lukumäärä", 0, 0),
        (None, None, None, None, "Keskiarvo", 0, 0),
        (None, None, None, None, "Maksimi", 0, 0),
        (None, None, None, None, "Minimi", 0, 0),
        (None, None, None, None, None, None, None),
        (None, None, None, "Kalleusalue 3", "Lukumäärä", 0, 0),
        (None, None, None, None, "Keskiarvo", 0, 0),
        (None, None, None, None, "Maksimi", 0, 0),
        (None, None, None, None, "Minimi", 0, 0),
        (None, None, None, None, None, None, None),
        (None, None, None, "Kalleusalue 4", "Lukumäärä", 0, 0),
        (None, None, None, None, "Keskiarvo", 0, 0),
        (None, None, None, None, "Maksimi", 0, 0),
        (None, None, None, None, "Minimi", 0, 0),
    ]


@pytest.mark.django_db
def test__api__sales_report__multiple_sales__one_excluded_from_sales(api_client: HitasAPIClient):
    # Create sales in the report interval
    sale_1: ApartmentSale = ApartmentSaleFactory.create(
        purchase_date=datetime.date(2020, 1, 1),
        purchase_price=50_000,
        apartment_share_of_housing_company_loans=10_000,
        apartment__surface_area=100,
        apartment__building__real_estate__housing_company__postal_code__value="00001",
        apartment__building__real_estate__housing_company__postal_code__cost_area=1,
    )
    # Sale not included even though in the report interval because of exclude_from_statistics
    ApartmentSaleFactory.create(
        purchase_date=datetime.date(2020, 1, 15),
        purchase_price=130_000,
        exclude_from_statistics=True,
        apartment_share_of_housing_company_loans=100,
        apartment__surface_area=100,
        apartment__building__real_estate__housing_company__postal_code__value="00002",
        apartment__building__real_estate__housing_company__postal_code__cost_area=1,
    )

    data = {
        "start_date": "2020-01-01",
        "end_date": "2020-01-31",
    }
    url = reverse("hitas:sales-report-list") + "?" + urlencode(data)
    response: HttpResponse = api_client.get(url)

    assert response.status_code == status.HTTP_200_OK

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert list(worksheet.values) == [
        (
            "Kalleusalue",
            "Postinumero",
            "Osoite",
            "Ilmoituspäivä",
            "Kauppapäivä",
            "Kaupan neliöhinta",
            "Velaton neliöhinta",
        ),
        (
            sale_1.apartment.postal_code.cost_area,
            sale_1.apartment.postal_code.value,
            sale_1.apartment.address,
            datetime.datetime.fromisoformat(sale_1.notification_date.isoformat()),
            datetime.datetime.fromisoformat(sale_1.purchase_date.isoformat()),
            500,  # 50_000 / 100
            600,  # (50_000 + 10_000) / 100
        ),
        (None, None, None, None, None, None, None),
        (None, None, None, "Kaikki kaupat", "Lukumäärä", 1, 1),
        (None, None, None, None, "Keskiarvo", 500, 600),
        (None, None, None, None, "Maksimi", 500, 600),
        (None, None, None, None, "Minimi", 500, 600),
        (None, None, None, None, None, None, None),
        (None, None, None, "Kalleusalue 1", "Lukumäärä", 1, 1),
        (None, None, None, None, "Keskiarvo", 500, 600),
        (None, None, None, None, "Maksimi", 500, 600),
        (None, None, None, None, "Minimi", 500, 600),
        (None, None, None, None, None, None, None),
        (None, None, None, "Kalleusalue 2", "Lukumäärä", 0, 0),
        (None, None, None, None, "Keskiarvo", 0, 0),
        (None, None, None, None, "Maksimi", 0, 0),
        (None, None, None, None, "Minimi", 0, 0),
        (None, None, None, None, None, None, None),
        (None, None, None, "Kalleusalue 3", "Lukumäärä", 0, 0),
        (None, None, None, None, "Keskiarvo", 0, 0),
        (None, None, None, None, "Maksimi", 0, 0),
        (None, None, None, None, "Minimi", 0, 0),
        (None, None, None, None, None, None, None),
        (None, None, None, "Kalleusalue 4", "Lukumäärä", 0, 0),
        (None, None, None, None, "Keskiarvo", 0, 0),
        (None, None, None, None, "Maksimi", 0, 0),
        (None, None, None, None, "Minimi", 0, 0),
    ]


@pytest.mark.django_db
def test__api__sales_report__multiple_sales__one_out_of_interval(api_client: HitasAPIClient):
    # Create sales in the report interval
    sale: ApartmentSale = ApartmentSaleFactory.create(
        purchase_date=datetime.date(2020, 1, 1),
        purchase_price=50_000,
        apartment_share_of_housing_company_loans=10_000,
        apartment__surface_area=100,
        apartment__building__real_estate__housing_company__postal_code__value="00001",
        apartment__building__real_estate__housing_company__postal_code__cost_area=1,
    )
    ApartmentSaleFactory.create(
        purchase_date=datetime.date(2020, 2, 1),
        purchase_price=130_000,
        apartment_share_of_housing_company_loans=100,
        apartment__surface_area=100,
        apartment__building__real_estate__housing_company__postal_code__value="00002",
        apartment__building__real_estate__housing_company__postal_code__cost_area=1,
    )

    data = {
        "start_date": "2020-01-01",
        "end_date": "2020-01-31",
    }
    url = reverse("hitas:sales-report-list") + "?" + urlencode(data)
    response: HttpResponse = api_client.get(url)

    assert response.status_code == status.HTTP_200_OK

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert list(worksheet.values) == [
        (
            "Kalleusalue",
            "Postinumero",
            "Osoite",
            "Ilmoituspäivä",
            "Kauppapäivä",
            "Kaupan neliöhinta",
            "Velaton neliöhinta",
        ),
        (
            sale.apartment.postal_code.cost_area,
            sale.apartment.postal_code.value,
            sale.apartment.address,
            datetime.datetime.fromisoformat(sale.notification_date.isoformat()),
            datetime.datetime.fromisoformat(sale.purchase_date.isoformat()),
            500,  # 50_000 / 100
            600,  # (50_000 + 10_000) / 100
        ),
        (None, None, None, None, None, None, None),
        (None, None, None, "Kaikki kaupat", "Lukumäärä", 1, 1),
        (None, None, None, None, "Keskiarvo", 500, 600),
        (None, None, None, None, "Maksimi", 500, 600),
        (None, None, None, None, "Minimi", 500, 600),
        (None, None, None, None, None, None, None),
        (None, None, None, "Kalleusalue 1", "Lukumäärä", 1, 1),
        (None, None, None, None, "Keskiarvo", 500, 600),
        (None, None, None, None, "Maksimi", 500, 600),
        (None, None, None, None, "Minimi", 500, 600),
        (None, None, None, None, None, None, None),
        (None, None, None, "Kalleusalue 2", "Lukumäärä", 0, 0),
        (None, None, None, None, "Keskiarvo", 0, 0),
        (None, None, None, None, "Maksimi", 0, 0),
        (None, None, None, None, "Minimi", 0, 0),
        (None, None, None, None, None, None, None),
        (None, None, None, "Kalleusalue 3", "Lukumäärä", 0, 0),
        (None, None, None, None, "Keskiarvo", 0, 0),
        (None, None, None, None, "Maksimi", 0, 0),
        (None, None, None, None, "Minimi", 0, 0),
        (None, None, None, None, None, None, None),
        (None, None, None, "Kalleusalue 4", "Lukumäärä", 0, 0),
        (None, None, None, None, "Keskiarvo", 0, 0),
        (None, None, None, None, "Maksimi", 0, 0),
        (None, None, None, None, "Minimi", 0, 0),
    ]


@pytest.mark.django_db
def test__api__sales_report__start_date_after_end_date(api_client: HitasAPIClient):
    # Create sales in the report interval
    ApartmentSaleFactory.create(
        purchase_date=datetime.date(2020, 1, 1),
        purchase_price=50_000,
        apartment_share_of_housing_company_loans=10_000,
        apartment__surface_area=100,
        apartment__building__real_estate__housing_company__postal_code__value="00001",
        apartment__building__real_estate__housing_company__postal_code__cost_area=1,
    )

    data = {
        "start_date": "2020-01-31",
        "end_date": "2020-01-01",
    }
    url = reverse("hitas:sales-report-list") + "?" + urlencode(data)
    response = api_client.get(url)

    assert response.status_code == status.HTTP_400_BAD_REQUEST
    assert response.json() == {
        "error": "bad_request",
        "fields": [
            {
                "field": "non_field_errors",
                "message": "Start date must be before end date",
            },
        ],
        "message": "Bad request",
        "reason": "Bad Request",
        "status": 400,
    }


@pytest.mark.django_db
def test__api__sales_report__surface_area_missing(api_client: HitasAPIClient):
    # Create sales in the report interval
    sale: ApartmentSale = ApartmentSaleFactory.create(
        purchase_date=datetime.date(2020, 1, 1),
        purchase_price=50_000,
        apartment_share_of_housing_company_loans=10_000,
        apartment__surface_area=None,
        apartment__building__real_estate__housing_company__postal_code__value="00001",
        apartment__building__real_estate__housing_company__postal_code__cost_area=1,
    )

    data = {
        "start_date": "2020-01-01",
        "end_date": "2020-01-31",
    }
    url = reverse("hitas:sales-report-list") + "?" + urlencode(data)
    response = api_client.get(url)

    assert response.status_code == status.HTTP_400_BAD_REQUEST
    assert response.json() == {
        "error": "bad_request",
        "fields": [
            {
                "field": "non_field_errors",
                "message": (
                    f"Surface are zero or missing for apartment {sale.apartment.address!r}. "
                    f"Cannot calculate price per square meter."
                ),
            }
        ],
        "message": "Bad request",
        "reason": "Bad Request",
        "status": 400,
    }


# Regulated housing companies report


@pytest.mark.django_db
def test__api__regulated_housing_companies_report__single_housing_company(api_client: HitasAPIClient):
    housing_company: HousingCompany = HousingCompanyFactory.create(
        postal_code__value="00001",
        postal_code__cost_area=1,
        hitas_type=HitasType.HITAS_I,
    )
    sale: ApartmentSale = ApartmentSaleFactory.create(
        purchase_date=datetime.date(2020, 1, 1),
        purchase_price=50_000,
        apartment_share_of_housing_company_loans=10_000,
        apartment__surface_area=100,
        apartment__completion_date=datetime.date(2020, 1, 1),
        apartment__building__real_estate__housing_company=housing_company,
    )

    url = reverse("hitas:regulated-housing-companies-report-list")
    response: HttpResponse = api_client.get(url)

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert list(worksheet.values) == [
        (
            "Kalleusalue",
            "Postinumero",
            "Yhtiö",
            "Osoite",
            "Valmistumispäivä",
            "Asuntojen lukumäärä",
            "Keskineliöhinta",
        ),
        (
            housing_company.postal_code.cost_area,
            housing_company.postal_code.value,
            housing_company.display_name,
            housing_company.street_address,
            datetime.datetime.fromisoformat(sale.apartment.completion_date.isoformat()),
            1,
            int(sale.total_price / sale.apartment.surface_area),
        ),
        (None, None, None, None, None, None, None),  # Empty row at the bottom for filtering and sorting
    ]


@pytest.mark.django_db
def test__api__regulated_housing_companies_report__multiple_housing_companies(api_client: HitasAPIClient):
    housing_company_1: HousingCompany = HousingCompanyFactory.create(
        postal_code__value="00001",
        postal_code__cost_area=1,
        hitas_type=HitasType.HITAS_I,
    )
    sale_1: ApartmentSale = ApartmentSaleFactory.create(
        purchase_date=datetime.date(2020, 1, 1),
        purchase_price=50_000,
        apartment_share_of_housing_company_loans=10_000,
        apartment__surface_area=100,
        apartment__completion_date=datetime.date(2020, 1, 1),
        apartment__building__real_estate__housing_company=housing_company_1,
    )

    housing_company_2: HousingCompany = HousingCompanyFactory.create(
        postal_code__value="00002",
        postal_code__cost_area=1,
        hitas_type=HitasType.HITAS_II,
    )
    sale_2: ApartmentSale = ApartmentSaleFactory.create(
        purchase_date=datetime.date(2021, 1, 1),
        purchase_price=100_000,
        apartment_share_of_housing_company_loans=44_000,
        apartment__surface_area=50,
        apartment__completion_date=datetime.date(2021, 1, 1),
        apartment__building__real_estate__housing_company=housing_company_2,
    )

    url = reverse("hitas:regulated-housing-companies-report-list")
    response: HttpResponse = api_client.get(url)

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert list(worksheet.values) == [
        (
            "Kalleusalue",
            "Postinumero",
            "Yhtiö",
            "Osoite",
            "Valmistumispäivä",
            "Asuntojen lukumäärä",
            "Keskineliöhinta",
        ),
        (
            housing_company_1.postal_code.cost_area,
            housing_company_1.postal_code.value,
            housing_company_1.display_name,
            housing_company_1.street_address,
            datetime.datetime.fromisoformat(sale_1.apartment.completion_date.isoformat()),
            1,
            int(sale_1.total_price / sale_1.apartment.surface_area),
        ),
        (
            housing_company_2.postal_code.cost_area,
            housing_company_2.postal_code.value,
            housing_company_2.display_name,
            housing_company_2.street_address,
            datetime.datetime.fromisoformat(sale_2.apartment.completion_date.isoformat()),
            1,
            int(sale_2.total_price / sale_2.apartment.surface_area),
        ),
        (None, None, None, None, None, None, None),  # Empty row at the bottom for filtering and sorting
    ]


@pytest.mark.django_db
def test__api__half_hitas_housing_companies_report__multiple_housing_companies(api_client: HitasAPIClient):
    housing_company_1: HousingCompany = HousingCompanyFactory.create(
        postal_code__value="00001",
        postal_code__cost_area=1,
        hitas_type=HitasType.HALF_HITAS,
    )
    sale_1: ApartmentSale = ApartmentSaleFactory.create(
        purchase_date=datetime.date(2020, 1, 1),
        purchase_price=50_000,
        apartment_share_of_housing_company_loans=10_000,
        apartment__surface_area=100,
        apartment__completion_date=datetime.date(2020, 1, 1),
        apartment__building__real_estate__housing_company=housing_company_1,
    )

    housing_company_2: HousingCompany = HousingCompanyFactory.create(
        postal_code__value="00002",
        postal_code__cost_area=1,
        hitas_type=HitasType.HALF_HITAS,
    )
    sale_2: ApartmentSale = ApartmentSaleFactory.create(
        purchase_date=datetime.date(2021, 1, 1),
        purchase_price=100_000,
        apartment_share_of_housing_company_loans=44_000,
        apartment__surface_area=50,
        apartment__completion_date=datetime.date(2021, 1, 1),
        apartment__building__real_estate__housing_company=housing_company_2,
    )

    url = reverse("hitas:half-hitas-housing-companies-report-list")
    response: HttpResponse = api_client.get(url)

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert list(worksheet.values) == [
        (
            "Kalleusalue",
            "Postinumero",
            "Yhtiö",
            "Osoite",
            "Valmistumispäivä",
            "Asuntojen lukumäärä",
            "Keskineliöhinta",
        ),
        (
            housing_company_1.postal_code.cost_area,
            housing_company_1.postal_code.value,
            housing_company_1.display_name,
            housing_company_1.street_address,
            datetime.datetime.fromisoformat(sale_1.apartment.completion_date.isoformat()),
            1,
            int(sale_1.total_price / sale_1.apartment.surface_area),
        ),
        (
            housing_company_2.postal_code.cost_area,
            housing_company_2.postal_code.value,
            housing_company_2.display_name,
            housing_company_2.street_address,
            datetime.datetime.fromisoformat(sale_2.apartment.completion_date.isoformat()),
            1,
            int(sale_2.total_price / sale_2.apartment.surface_area),
        ),
        (None, None, None, None, None, None, None),  # Empty row at the bottom for filtering and sorting
    ]


@pytest.mark.django_db
def test__api__regulated_housing_companies_report__unregulated_not_included(api_client: HitasAPIClient):
    housing_company_1: HousingCompany = HousingCompanyFactory.create(
        postal_code__value="00001",
        postal_code__cost_area=1,
        hitas_type=HitasType.HITAS_I,
    )
    sale_1: ApartmentSale = ApartmentSaleFactory.create(
        purchase_date=datetime.date(2020, 1, 1),
        purchase_price=50_000,
        apartment_share_of_housing_company_loans=10_000,
        apartment__surface_area=100,
        apartment__completion_date=datetime.date(2020, 1, 1),
        apartment__building__real_estate__housing_company=housing_company_1,
    )

    # Housing company not included in the report since it's released from regulation
    housing_company_2: HousingCompany = HousingCompanyFactory.create(
        postal_code__value="00002",
        postal_code__cost_area=1,
        regulation_status=RegulationStatus.RELEASED_BY_HITAS,
        hitas_type=HitasType.HITAS_I,
    )
    ApartmentSaleFactory.create(
        purchase_date=datetime.date(2021, 1, 1),
        purchase_price=100_000,
        apartment_share_of_housing_company_loans=44_000,
        apartment__surface_area=50,
        apartment__completion_date=datetime.date(2021, 1, 1),
        apartment__building__real_estate__housing_company=housing_company_2,
    )

    # Housing company not included in the report since HALF_HITAS have a separate report
    housing_company_3: HousingCompany = HousingCompanyFactory.create(
        postal_code__value="00001",
        postal_code__cost_area=1,
        hitas_type=HitasType.HALF_HITAS,
    )
    ApartmentSaleFactory.create(
        purchase_date=datetime.date(2020, 1, 1),
        purchase_price=50_000,
        apartment_share_of_housing_company_loans=10_000,
        apartment__surface_area=100,
        apartment__completion_date=datetime.date(2020, 1, 1),
        apartment__building__real_estate__housing_company=housing_company_3,
    )

    url = reverse("hitas:regulated-housing-companies-report-list")
    response: HttpResponse = api_client.get(url)

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert list(worksheet.values) == [
        (
            "Kalleusalue",
            "Postinumero",
            "Yhtiö",
            "Osoite",
            "Valmistumispäivä",
            "Asuntojen lukumäärä",
            "Keskineliöhinta",
        ),
        (
            housing_company_1.postal_code.cost_area,
            housing_company_1.postal_code.value,
            housing_company_1.display_name,
            housing_company_1.street_address,
            datetime.datetime.fromisoformat(sale_1.apartment.completion_date.isoformat()),
            1,
            int(sale_1.total_price / sale_1.apartment.surface_area),
        ),
        (None, None, None, None, None, None, None),  # Empty row at the bottom for filtering and sorting
    ]


# Unregulated housing companies report


@pytest.mark.django_db
def test__api__unregulated_housing_companies_report__legacy_release_date(api_client: HitasAPIClient):
    housing_company: HousingCompany = HousingCompanyFactory.create(
        postal_code__value="00001",
        regulation_status=RegulationStatus.RELEASED_BY_HITAS,
        legacy_release_date=datetime.date(2022, 1, 1),
    )
    ApartmentFactory.create(
        completion_date=datetime.date(2020, 1, 1),
        building__real_estate__housing_company=housing_company,
    )
    apartment_2: Apartment = ApartmentFactory.create(
        completion_date=datetime.date(2021, 1, 1),
        building__real_estate__housing_company=housing_company,
    )

    url = reverse("hitas:unregulated-housing-companies-report-list")
    response: HttpResponse = api_client.get(url)

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert list(worksheet.values) == [
        (
            "Yhtiö",
            "Postinumero",
            "Valmistumispäivä",
            "Vapautumispäivä",
            "Vapautettu tontit yksikön toimesta?",
            "Asuntojen lukumäärä",
        ),
        (
            housing_company.display_name,
            housing_company.postal_code.value,
            datetime.datetime.fromisoformat(apartment_2.completion_date.isoformat()),
            datetime.datetime.fromisoformat(housing_company.legacy_release_date.isoformat()),
            None,
            2,
        ),
        (None, None, None, None, None, None),
        ("Taloyhtiötä yhteensä", 1, None, None, None, None),
        ("Asuntoja yhteensä", 2, None, None, None, None),
    ]


@pytest.mark.django_db
def test__api__unregulated_housing_companies_report__regulation_release_date(api_client: HitasAPIClient):
    housing_company: HousingCompany = HousingCompanyFactory.create(
        postal_code__value="00001",
        regulation_status=RegulationStatus.RELEASED_BY_HITAS,
    )
    apartment_1: Apartment = ApartmentFactory.create(
        completion_date=datetime.date(2020, 1, 1),
        building__real_estate__housing_company=housing_company,
    )
    apartment_2: Apartment = ApartmentFactory.create(
        completion_date=datetime.date(2021, 1, 1),
        building__real_estate__housing_company=housing_company,
    )

    results = ThirtyYearRegulationResults.objects.create(
        calculation_month=datetime.date(2001, 1, 1),
        regulation_month=datetime.date(2000, 1, 1),
        surface_area_price_ceiling=None,
        sales_data=FullSalesData(internal={}, external={}, price_by_area={}),
        replacement_postal_codes=[],
    )

    ThirtyYearRegulationResultsRow.objects.create(
        parent=results,
        housing_company=housing_company,
        completion_date=apartment_2.completion_date,
        surface_area=apartment_1.surface_area + apartment_2.surface_area,
        postal_code=housing_company.postal_code.value,
        realized_acquisition_price=100,
        unadjusted_average_price_per_square_meter=200,
        adjusted_average_price_per_square_meter=300,
        completion_month_index=400,
        calculation_month_index=500,
        regulation_result=RegulationResult.RELEASED_FROM_REGULATION,
        letter_fetched=False,
    )

    url = reverse("hitas:unregulated-housing-companies-report-list")
    response: HttpResponse = api_client.get(url)

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert list(worksheet.values) == [
        (
            "Yhtiö",
            "Postinumero",
            "Valmistumispäivä",
            "Vapautumispäivä",
            "Vapautettu tontit yksikön toimesta?",
            "Asuntojen lukumäärä",
        ),
        (
            housing_company.display_name,
            housing_company.postal_code.value,
            datetime.datetime.fromisoformat(apartment_2.completion_date.isoformat()),
            datetime.datetime.fromisoformat(results.calculation_month.isoformat()),
            None,
            2,
        ),
        (None, None, None, None, None, None),
        ("Taloyhtiötä yhteensä", 1, None, None, None, None),
        ("Asuntoja yhteensä", 2, None, None, None, None),
    ]


@pytest.mark.django_db
def test__api__unregulated_housing_companies_report__released_by_plot_department(api_client: HitasAPIClient):
    housing_company: HousingCompany = HousingCompanyFactory.create(
        postal_code__value="00001",
        regulation_status=RegulationStatus.RELEASED_BY_PLOT_DEPARTMENT,
        legacy_release_date=datetime.date(2022, 1, 1),
    )
    apartment: Apartment = ApartmentFactory.create(
        completion_date=datetime.date(2020, 1, 1),
        building__real_estate__housing_company=housing_company,
    )

    url = reverse("hitas:unregulated-housing-companies-report-list")
    response: HttpResponse = api_client.get(url)

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert list(worksheet.values) == [
        (
            "Yhtiö",
            "Postinumero",
            "Valmistumispäivä",
            "Vapautumispäivä",
            "Vapautettu tontit yksikön toimesta?",
            "Asuntojen lukumäärä",
        ),
        (
            housing_company.display_name,
            housing_company.postal_code.value,
            datetime.datetime.fromisoformat(apartment.completion_date.isoformat()),
            datetime.datetime.fromisoformat(housing_company.legacy_release_date.isoformat()),
            "X",
            1,
        ),
        (None, None, None, None, None, None),
        ("Taloyhtiötä yhteensä", 1, None, None, None, None),
        ("Asuntoja yhteensä", 1, None, None, None, None),
    ]


@pytest.mark.django_db
def test__api__unregulated_housing_companies_report__multiple_housing_companies(api_client: HitasAPIClient):
    housing_company_1: HousingCompany = HousingCompanyFactory.create(
        postal_code__value="00001",
        regulation_status=RegulationStatus.RELEASED_BY_HITAS,
        legacy_release_date=datetime.date(2022, 1, 1),
    )
    ApartmentFactory.create(
        completion_date=datetime.date(2020, 1, 1),
        building__real_estate__housing_company=housing_company_1,
    )
    apartment_2: Apartment = ApartmentFactory.create(
        completion_date=datetime.date(2021, 1, 1),
        building__real_estate__housing_company=housing_company_1,
    )

    housing_company_2: HousingCompany = HousingCompanyFactory.create(
        postal_code__value="00001",
        regulation_status=RegulationStatus.RELEASED_BY_PLOT_DEPARTMENT,
        legacy_release_date=datetime.date(2004, 1, 1),
    )
    apartment_3: Apartment = ApartmentFactory.create(
        completion_date=datetime.date(2000, 1, 1),
        building__real_estate__housing_company=housing_company_2,
    )

    url = reverse("hitas:unregulated-housing-companies-report-list")
    response: HttpResponse = api_client.get(url)

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert list(worksheet.values) == [
        (
            "Yhtiö",
            "Postinumero",
            "Valmistumispäivä",
            "Vapautumispäivä",
            "Vapautettu tontit yksikön toimesta?",
            "Asuntojen lukumäärä",
        ),
        (
            housing_company_1.display_name,
            housing_company_1.postal_code.value,
            datetime.datetime.fromisoformat(apartment_2.completion_date.isoformat()),
            datetime.datetime.fromisoformat(housing_company_1.legacy_release_date.isoformat()),
            None,
            2,
        ),
        (
            housing_company_2.display_name,
            housing_company_2.postal_code.value,
            datetime.datetime.fromisoformat(apartment_3.completion_date.isoformat()),
            datetime.datetime.fromisoformat(housing_company_2.legacy_release_date.isoformat()),
            "X",
            1,
        ),
        (None, None, None, None, None, None),
        ("Taloyhtiötä yhteensä", 2, None, None, None, None),
        ("Asuntoja yhteensä", 3, None, None, None, None),
    ]


# Housing company state report


@pytest.mark.django_db
def test__api__housing_company_states_report__not_completed(api_client: HitasAPIClient):
    housing_company: HousingCompany = HousingCompanyFactory.create(
        regulation_status=RegulationStatus.REGULATED,
        hitas_type=HitasType.NEW_HITAS_I,
    )
    ApartmentFactory.create(
        completion_date=datetime.date(2020, 1, 1),
        building__real_estate__housing_company=housing_company,
    )
    # This apartment is not ready, so housing company is not ready either
    ApartmentFactory.create(
        completion_date=None,
        building__real_estate__housing_company=housing_company,
    )

    url = reverse("hitas:housing-company-states-report-list")
    response: HttpResponse = api_client.get(url)

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert list(worksheet.values) == [
        ("Taloyhtiön tila", "Taloyhtiöiden lukumäärä", "Asuntojen lukumäärä"),
        ("Ei valmis", 1, 2),
        ("Sääntelyn piirissä", 0, 0),
        ("Sääntelystä vapautuneet", 0, 0),
        ("Vapautuneet tontit-yksikön päätöksellä", 0, 0),
        ("Puolihitas yhtiöt", 0, 0),
        (None, None, None),
        ("Yhtiöitä yhteensä", 1, None),
        ("Asuntoja yhteensä", 2, None),
    ]


@pytest.mark.django_db
@pytest.mark.parametrize(
    "status",
    [
        RegulationStatus.REGULATED,
        RegulationStatus.RELEASED_BY_HITAS,
        RegulationStatus.RELEASED_BY_PLOT_DEPARTMENT,
    ],
)
def test__api__housing_company_states_report__regulation_status(api_client: HitasAPIClient, status):
    housing_company: HousingCompany = HousingCompanyFactory.create(
        regulation_status=status,
        hitas_type=HitasType.NEW_HITAS_I,
    )
    ApartmentFactory.create(
        completion_date=datetime.date(2020, 1, 1),
        building__real_estate__housing_company=housing_company,
    )
    ApartmentFactory.create(
        completion_date=datetime.date(2020, 1, 1),
        building__real_estate__housing_company=housing_company,
    )

    url = reverse("hitas:housing-company-states-report-list")
    response: HttpResponse = api_client.get(url)

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert list(worksheet.values) == [
        ("Taloyhtiön tila", "Taloyhtiöiden lukumäärä", "Asuntojen lukumäärä"),
        ("Ei valmis", 0, 0),
        (
            "Sääntelyn piirissä",
            1 if status == RegulationStatus.REGULATED else 0,
            2 if status == RegulationStatus.REGULATED else 0,
        ),
        (
            "Sääntelystä vapautuneet",
            1 if status == RegulationStatus.RELEASED_BY_HITAS else 0,
            2 if status == RegulationStatus.RELEASED_BY_HITAS else 0,
        ),
        (
            "Vapautuneet tontit-yksikön päätöksellä",
            1 if status == RegulationStatus.RELEASED_BY_PLOT_DEPARTMENT else 0,
            2 if status == RegulationStatus.RELEASED_BY_PLOT_DEPARTMENT else 0,
        ),
        ("Puolihitas yhtiöt", 0, 0),
        (None, None, None),
        ("Yhtiöitä yhteensä", 1, None),
        ("Asuntoja yhteensä", 2, None),
    ]


@pytest.mark.django_db
def test__api__housing_company_states_report__half_hitas(api_client: HitasAPIClient):
    housing_company: HousingCompany = HousingCompanyFactory.create(
        regulation_status=RegulationStatus.REGULATED,
        hitas_type=HitasType.HALF_HITAS,
    )
    ApartmentFactory.create(
        completion_date=datetime.date(2020, 1, 1),
        building__real_estate__housing_company=housing_company,
    )
    ApartmentFactory.create(
        completion_date=datetime.date(2020, 1, 1),
        building__real_estate__housing_company=housing_company,
    )

    url = reverse("hitas:housing-company-states-report-list")
    response: HttpResponse = api_client.get(url)

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert list(worksheet.values) == [
        ("Taloyhtiön tila", "Taloyhtiöiden lukumäärä", "Asuntojen lukumäärä"),
        ("Ei valmis", 0, 0),
        ("Sääntelyn piirissä", 0, 0),
        ("Sääntelystä vapautuneet", 0, 0),
        ("Vapautuneet tontit-yksikön päätöksellä", 0, 0),
        ("Puolihitas yhtiöt", 1, 2),
        (None, None, None),
        ("Yhtiöitä yhteensä", 1, None),
        ("Asuntoja yhteensä", 2, None),
    ]


@pytest.mark.django_db
def test__api__housing_company_states_report__one_of_each(api_client: HitasAPIClient):
    housing_company: HousingCompany = HousingCompanyFactory.create(
        regulation_status=RegulationStatus.REGULATED,
        hitas_type=HitasType.NEW_HITAS_I,
    )
    ApartmentFactory.create(
        completion_date=None,
        building__real_estate__housing_company=housing_company,
    )

    housing_company: HousingCompany = HousingCompanyFactory.create(
        regulation_status=RegulationStatus.REGULATED,
        hitas_type=HitasType.HALF_HITAS,
    )
    ApartmentFactory.create(
        completion_date=datetime.date(2020, 1, 1),
        building__real_estate__housing_company=housing_company,
    )

    housing_company: HousingCompany = HousingCompanyFactory.create(
        regulation_status=RegulationStatus.REGULATED,
        hitas_type=HitasType.NEW_HITAS_I,
    )
    ApartmentFactory.create(
        completion_date=datetime.date(2020, 1, 1),
        building__real_estate__housing_company=housing_company,
    )

    housing_company: HousingCompany = HousingCompanyFactory.create(
        regulation_status=RegulationStatus.RELEASED_BY_HITAS,
        hitas_type=HitasType.NEW_HITAS_I,
    )
    ApartmentFactory.create(
        completion_date=datetime.date(2020, 1, 1),
        building__real_estate__housing_company=housing_company,
    )

    housing_company: HousingCompany = HousingCompanyFactory.create(
        regulation_status=RegulationStatus.RELEASED_BY_PLOT_DEPARTMENT,
        hitas_type=HitasType.NEW_HITAS_I,
    )
    ApartmentFactory.create(
        completion_date=datetime.date(2020, 1, 1),
        building__real_estate__housing_company=housing_company,
    )

    url = reverse("hitas:housing-company-states-report-list")
    response: HttpResponse = api_client.get(url)

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert list(worksheet.values) == [
        ("Taloyhtiön tila", "Taloyhtiöiden lukumäärä", "Asuntojen lukumäärä"),
        ("Ei valmis", 1, 1),
        ("Sääntelyn piirissä", 1, 1),
        ("Sääntelystä vapautuneet", 1, 1),
        ("Vapautuneet tontit-yksikön päätöksellä", 1, 1),
        ("Puolihitas yhtiöt", 1, 1),
        (None, None, None),
        ("Yhtiöitä yhteensä", 5, None),
        ("Asuntoja yhteensä", 5, None),
    ]


@pytest.mark.django_db
def test__api__housing_company_states_report__json_endpoint(api_client: HitasAPIClient):
    housing_company: HousingCompany = HousingCompanyFactory.create(
        regulation_status=RegulationStatus.REGULATED,
        hitas_type=HitasType.NEW_HITAS_I,
    )
    ApartmentFactory.create(
        completion_date=None,
        building__real_estate__housing_company=housing_company,
    )

    housing_company: HousingCompany = HousingCompanyFactory.create(
        regulation_status=RegulationStatus.REGULATED,
        hitas_type=HitasType.HALF_HITAS,
    )
    ApartmentFactory.create(
        completion_date=datetime.date(2020, 1, 1),
        building__real_estate__housing_company=housing_company,
    )

    housing_company: HousingCompany = HousingCompanyFactory.create(
        regulation_status=RegulationStatus.REGULATED,
        hitas_type=HitasType.NEW_HITAS_I,
    )
    ApartmentFactory.create(
        completion_date=datetime.date(2020, 1, 1),
        building__real_estate__housing_company=housing_company,
    )

    housing_company: HousingCompany = HousingCompanyFactory.create(
        regulation_status=RegulationStatus.RELEASED_BY_HITAS,
        hitas_type=HitasType.NEW_HITAS_I,
    )
    ApartmentFactory.create(
        completion_date=datetime.date(2020, 1, 1),
        building__real_estate__housing_company=housing_company,
    )

    housing_company: HousingCompany = HousingCompanyFactory.create(
        regulation_status=RegulationStatus.RELEASED_BY_PLOT_DEPARTMENT,
        hitas_type=HitasType.NEW_HITAS_I,
    )
    ApartmentFactory.create(
        completion_date=datetime.date(2020, 1, 1),
        building__real_estate__housing_company=housing_company,
    )

    url = reverse("hitas:housing-company-states-list")
    response = api_client.get(url)

    assert response.status_code == status.HTTP_200_OK, response.json()
    assert response.json() == [
        {
            "status": "Ei valmis",
            "housing_company_count": 1,
            "apartment_count": 1,
        },
        {
            "status": "Sääntelyn piirissä",
            "housing_company_count": 1,
            "apartment_count": 1,
        },
        {
            "status": "Sääntelystä vapautuneet",
            "housing_company_count": 1,
            "apartment_count": 1,
        },
        {
            "status": "Vapautuneet tontit-yksikön päätöksellä",
            "housing_company_count": 1,
            "apartment_count": 1,
        },
        {
            "status": "Puolihitas yhtiöt",
            "housing_company_count": 1,
            "apartment_count": 1,
        },
    ]


# Sales by area report


@pytest.mark.django_db
def test__api__sales_by_area_report__no_surface_area(api_client: HitasAPIClient):
    sale: ApartmentSale = ApartmentSaleFactory.create(
        purchase_date=datetime.date(2020, 1, 1),
        purchase_price=50_000,
        apartment_share_of_housing_company_loans=10_000,
        apartment__surface_area=None,
        apartment__rooms=1,
        apartment__building__real_estate__housing_company__postal_code__value="00001",
        apartment__building__real_estate__housing_company__postal_code__cost_area=1,
    )

    data = {
        "start_date": "2020-01-01",
        "end_date": "2020-01-31",
    }
    url = reverse("hitas:sales-by-postal-code-and-area-report-list") + "?" + urlencode(data)
    response = api_client.get(url)

    assert response.status_code == status.HTTP_400_BAD_REQUEST, response.json()
    assert response.json() == {
        "error": "bad_request",
        "fields": [
            {
                "field": "non_field_errors",
                "message": (
                    f"Surface are zero or missing for apartment {sale.apartment.address!r}. "
                    f"Cannot calculate price per square meter."
                ),
            }
        ],
        "message": "Bad request",
        "reason": "Bad Request",
        "status": 400,
    }


@pytest.mark.parametrize(
    ["rooms", "expected_label"],
    [
        [1, "1h"],
        [2, "2h"],
        [3, "3h+"],
        [4, "3h+"],
    ],
)
@pytest.mark.django_db
def test__api__sales_by_area_report__rooms(api_client: HitasAPIClient, rooms, expected_label):
    ApartmentSaleFactory.create(
        purchase_date=datetime.date(2020, 1, 1),
        purchase_price=50_000,
        apartment_share_of_housing_company_loans=10_000,
        apartment__surface_area=100,
        apartment__rooms=rooms,
        apartment__building__real_estate__housing_company__postal_code__value="00001",
        apartment__building__real_estate__housing_company__postal_code__cost_area=1,
    )

    data = {
        "start_date": "2020-01-01",
        "end_date": "2020-01-31",
    }
    url = reverse("hitas:sales-by-postal-code-and-area-report-list") + "?" + urlencode(data)
    response: HttpResponse = api_client.get(url)

    assert response.status_code == status.HTTP_200_OK

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert list(worksheet.values) == [
        (
            "Kalleusalue",
            "Postinumero",
            "Huoneluku",
            "Kauppojen lukumäärä",
            "Keskineliöhinta",
            "Alin neliöhinta",
            "Ylin neliöhinta",
        ),
        (1, "00001", expected_label, 1, 600, 600, 600),
        (None, None, None, None, None, None, None),
        (None, None, "Koko Helsingin alue", 1, 600, 600, 600),
    ]


@pytest.mark.django_db
def test__api__sales_by_area_report__multiple__same_postal_code__same_room_count(api_client: HitasAPIClient):
    ApartmentSaleFactory.create(
        purchase_date=datetime.date(2020, 1, 1),
        purchase_price=50_000,
        apartment_share_of_housing_company_loans=10_000,
        apartment__surface_area=100,
        apartment__rooms=1,
        apartment__building__real_estate__housing_company__postal_code__value="00001",
        apartment__building__real_estate__housing_company__postal_code__cost_area=1,
    )

    ApartmentSaleFactory.create(
        purchase_date=datetime.date(2020, 1, 1),
        purchase_price=20_000,
        apartment_share_of_housing_company_loans=5_000,
        apartment__surface_area=80,
        apartment__rooms=1,
        apartment__building__real_estate__housing_company__postal_code__value="00001",
        apartment__building__real_estate__housing_company__postal_code__cost_area=1,
    )

    data = {
        "start_date": "2020-01-01",
        "end_date": "2020-01-31",
    }
    url = reverse("hitas:sales-by-postal-code-and-area-report-list") + "?" + urlencode(data)
    response: HttpResponse = api_client.get(url)

    assert response.status_code == status.HTTP_200_OK

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert list(worksheet.values) == [
        (
            "Kalleusalue",
            "Postinumero",
            "Huoneluku",
            "Kauppojen lukumäärä",
            "Keskineliöhinta",
            "Alin neliöhinta",
            "Ylin neliöhinta",
        ),
        (1, "00001", "1h", 2, 456.25, 312.5, 600),
        (None, None, None, None, None, None, None),
        (None, None, "Koko Helsingin alue", 2, 456.25, 312.5, 600),
    ]


@pytest.mark.django_db
def test__api__sales_by_area_report__multiple__same_postal_code__different_room_count(api_client: HitasAPIClient):
    ApartmentSaleFactory.create(
        purchase_date=datetime.date(2020, 1, 1),
        purchase_price=50_000,
        apartment_share_of_housing_company_loans=10_000,
        apartment__surface_area=100,
        apartment__rooms=1,
        apartment__building__real_estate__housing_company__postal_code__value="00001",
        apartment__building__real_estate__housing_company__postal_code__cost_area=1,
    )

    ApartmentSaleFactory.create(
        purchase_date=datetime.date(2020, 1, 1),
        purchase_price=20_000,
        apartment_share_of_housing_company_loans=5_000,
        apartment__surface_area=80,
        apartment__rooms=2,
        apartment__building__real_estate__housing_company__postal_code__value="00001",
        apartment__building__real_estate__housing_company__postal_code__cost_area=1,
    )

    data = {
        "start_date": "2020-01-01",
        "end_date": "2020-01-31",
    }
    url = reverse("hitas:sales-by-postal-code-and-area-report-list") + "?" + urlencode(data)
    response: HttpResponse = api_client.get(url)

    assert response.status_code == status.HTTP_200_OK

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert list(worksheet.values) == [
        (
            "Kalleusalue",
            "Postinumero",
            "Huoneluku",
            "Kauppojen lukumäärä",
            "Keskineliöhinta",
            "Alin neliöhinta",
            "Ylin neliöhinta",
        ),
        (1, "00001", "1h", 1, 600, 600, 600),
        (1, "00001", "2h", 1, 312.5, 312.5, 312.5),
        (None, None, None, None, None, None, None),
        (None, None, "Koko Helsingin alue", 2, 456.25, 312.5, 600),
    ]


@pytest.mark.django_db
def test__api__sales_by_area_report__multiple__different_postal_code(api_client: HitasAPIClient):
    ApartmentSaleFactory.create(
        purchase_date=datetime.date(2020, 1, 1),
        purchase_price=50_000,
        apartment_share_of_housing_company_loans=10_000,
        apartment__surface_area=100,
        apartment__rooms=1,
        apartment__building__real_estate__housing_company__postal_code__value="00001",
        apartment__building__real_estate__housing_company__postal_code__cost_area=1,
    )

    ApartmentSaleFactory.create(
        purchase_date=datetime.date(2020, 1, 1),
        purchase_price=20_000,
        apartment_share_of_housing_company_loans=5_000,
        apartment__surface_area=80,
        apartment__rooms=1,
        apartment__building__real_estate__housing_company__postal_code__value="00002",
        apartment__building__real_estate__housing_company__postal_code__cost_area=1,
    )

    data = {
        "start_date": "2020-01-01",
        "end_date": "2020-01-31",
    }
    url = reverse("hitas:sales-by-postal-code-and-area-report-list") + "?" + urlencode(data)
    response: HttpResponse = api_client.get(url)

    assert response.status_code == status.HTTP_200_OK

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert list(worksheet.values) == [
        (
            "Kalleusalue",
            "Postinumero",
            "Huoneluku",
            "Kauppojen lukumäärä",
            "Keskineliöhinta",
            "Alin neliöhinta",
            "Ylin neliöhinta",
        ),
        (1, "00001", "1h", 1, 600, 600, 600),
        (1, "00002", "1h", 1, 312.5, 312.5, 312.5),
        (None, None, None, None, None, None, None),
        (None, None, "Koko Helsingin alue", 2, 456.25, 312.5, 600),
    ]


@pytest.mark.django_db
def test__api__sales_by_area_report__filter_resales(api_client: HitasAPIClient):
    # All apartments in the same area -> included in the same report row sale count
    area_args = {
        "apartment__building__real_estate__housing_company__postal_code__value": "00002",
        "apartment__building__real_estate__housing_company__postal_code__cost_area": 1,
    }
    sale_args = dict(**area_args, purchase_date=datetime.date(2020, 1, 1))
    ApartmentSaleFactory.create(**sale_args)  # Excluded: first sale
    apartment = ApartmentFactory.create(
        sales=[],
        building__real_estate__housing_company__postal_code__value="00002",
        building__real_estate__housing_company__postal_code__cost_area=1,
    )
    ApartmentSaleFactory.create(**sale_args, apartment=apartment)  # Excluded: first sale
    ApartmentSaleFactory.create(**sale_args, apartment=apartment)  # Included: resale
    sale_to_be_soft_deleted = ApartmentSaleFactory.create(**sale_args, apartment=apartment)  # Excluded: soft-deleted
    sale_to_be_soft_deleted.delete()
    ApartmentSaleFactory.create(**sale_args, apartment=apartment)  # Included: resale
    ApartmentSaleFactory.create(
        **area_args, purchase_date=datetime.date(2021, 1, 1), apartment=apartment
    )  # Excluded: outside of date range
    data = {
        "start_date": "2020-01-01",
        "end_date": "2020-01-31",
        "filter": "resale",
    }
    url = reverse("hitas:sales-by-postal-code-and-area-report-list") + "?" + urlencode(data)
    response: HttpResponse = api_client.get(url)
    assert response.status_code == status.HTTP_200_OK
    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]
    sales_count = list(worksheet.values)[1][3]
    assert sales_count == 2, "There should be two resales"


@pytest.mark.django_db
def test__api__sales_by_area_report__filter_firstsales(api_client: HitasAPIClient):
    # All apartments in the same area -> included in the same report row sale count
    area_args = {
        "apartment__building__real_estate__housing_company__postal_code__value": "00002",
        "apartment__building__real_estate__housing_company__postal_code__cost_area": 1,
    }
    sale_args = dict(**area_args, purchase_date=datetime.date(2020, 1, 1))
    ApartmentSaleFactory.create(**sale_args)  # Included: first sale
    sale_to_be_soft_deleted = ApartmentSaleFactory.create(**sale_args)  # Excluded: soft-deleted
    sale_to_be_soft_deleted.delete()
    ApartmentSaleFactory.create(**area_args, purchase_date=datetime.date(2021, 1, 1))  # Excluded: outside of date range
    apartment = ApartmentFactory.create(
        sales=[],
        building__real_estate__housing_company__postal_code__value="00002",
        building__real_estate__housing_company__postal_code__cost_area=1,
    )
    sale_to_be_soft_deleted = ApartmentSaleFactory.create(**sale_args, apartment=apartment)  # Excluded: soft-deleted
    sale_to_be_soft_deleted.delete()
    ApartmentSaleFactory.create(**sale_args, apartment=apartment)  # Included: first sale
    ApartmentSaleFactory.create(**sale_args, apartment=apartment)  # Excluded: resale
    data = {
        "start_date": "2020-01-01",
        "end_date": "2020-01-31",
        "filter": "firstsale",
    }
    url = reverse("hitas:sales-by-postal-code-and-area-report-list") + "?" + urlencode(data)
    response: HttpResponse = api_client.get(url)
    assert response.status_code == status.HTTP_200_OK
    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]
    sales_count = list(worksheet.values)[1][3]
    assert sales_count == 2, "There should be two first sales"


# Multiple ownerships report


@pytest.mark.django_db
def test__api__multiple_ownerships_report__no_owners(api_client: HitasAPIClient):
    url = reverse("hitas:multiple-ownerships-report-list")
    response: HttpResponse = api_client.get(url)

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert list(worksheet.values) == [
        (
            "Omistajan nimi",
            "Asunnon osoite",
            "Postinumero",
            "Omistajan asuntojen lukumäärä",
            "Omistajan henkilö- tai Y-tunnus",
            "Yhtiön nimi",
            "Yhtiön valmistumispäivä",
            "Kalleusalue",
        ),
    ]


@pytest.mark.django_db
def test__api__regulated_ownerships_report__no_owners(api_client: HitasAPIClient):
    url = reverse("hitas:regulated-ownerships-report-list")
    response: HttpResponse = api_client.get(url)

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert len(list(worksheet.values)) == 1, "There should be only the header row"


@pytest.mark.django_db
@pytest.mark.parametrize("non_disclosure", [False, True])
def test__api__multiple_ownerships_report__single_owner(api_client: HitasAPIClient, non_disclosure):
    owner: Owner = OwnerFactory.create(non_disclosure=non_disclosure)
    ownership_1: Ownership = OwnershipFactory.create(
        owner=owner,
        sale__apartment__building__real_estate__housing_company__postal_code__value="00001",
        sale__apartment__building__real_estate__housing_company__hitas_type=HitasType.NEW_HITAS_I,
    )
    ownership_2: Ownership = OwnershipFactory.create(
        owner=owner,
        sale__apartment__building__real_estate__housing_company__postal_code__value="00002",
        sale__apartment__building__real_estate__housing_company__hitas_type=HitasType.NEW_HITAS_I,
    )

    url = reverse("hitas:multiple-ownerships-report-list")
    response: HttpResponse = api_client.get(url)

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert list(worksheet.values) == [
        (
            "Omistajan nimi",
            "Asunnon osoite",
            "Postinumero",
            "Omistajan asuntojen lukumäärä",
            "Omistajan henkilö- tai Y-tunnus",
            "Yhtiön nimi",
            "Yhtiön valmistumispäivä",
            "Kalleusalue",
        ),
        (
            Owner.OBFUSCATED_OWNER_NAME if non_disclosure else ownership_1.owner.name,
            ownership_1.apartment.address,
            ownership_1.apartment.postal_code.value,
            2,
            None if non_disclosure else ownership_1.owner.identifier,
            ownership_1.apartment.building.real_estate.housing_company.display_name,
            datetime.datetime.combine(
                ownership_1.apartment.building.real_estate.housing_company.completion_date, datetime.datetime.min.time()
            ),
            ownership_1.apartment.postal_code.cost_area,
        ),
        (
            Owner.OBFUSCATED_OWNER_NAME if non_disclosure else ownership_2.owner.name,
            ownership_2.apartment.address,
            ownership_2.apartment.postal_code.value,
            2,
            None if non_disclosure else ownership_2.owner.identifier,
            ownership_2.apartment.building.real_estate.housing_company.display_name,
            datetime.datetime.combine(
                ownership_2.apartment.building.real_estate.housing_company.completion_date, datetime.datetime.min.time()
            ),
            ownership_2.apartment.postal_code.cost_area,
        ),
    ]


@pytest.mark.django_db
@pytest.mark.parametrize("non_disclosure", [False, True])
def test__api__regulated_ownerships_report__single_owner(api_client: HitasAPIClient, non_disclosure):
    owner: Owner = OwnerFactory.create(non_disclosure=non_disclosure)
    OwnershipFactory.create(
        owner=owner,
        sale__apartment__building__real_estate__housing_company__postal_code__value="00001",
        sale__apartment__building__real_estate__housing_company__hitas_type=HitasType.NEW_HITAS_I,
    )
    OwnershipFactory.create(
        owner=owner,
        sale__apartment__building__real_estate__housing_company__postal_code__value="00002",
        sale__apartment__building__real_estate__housing_company__hitas_type=HitasType.NEW_HITAS_I,
    )

    url = reverse("hitas:regulated-ownerships-report-list")
    response: HttpResponse = api_client.get(url)

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert len(list(worksheet.values)) == 3, "There should be 2 ownership rows and 1 header row"


@pytest.mark.django_db
def test__api__multiple_ownerships_report__multiple_owners(api_client: HitasAPIClient):
    owner_1: Owner = OwnerFactory.create(name="Owner 1")
    ownership_1: Ownership = OwnershipFactory.create(
        owner=owner_1,
        sale__apartment__building__real_estate__housing_company__postal_code__value="00001",
        sale__apartment__building__real_estate__housing_company__hitas_type=HitasType.NEW_HITAS_I,
    )
    ownership_2: Ownership = OwnershipFactory.create(
        owner=owner_1,
        sale__apartment__building__real_estate__housing_company__postal_code__value="00002",
        sale__apartment__building__real_estate__housing_company__hitas_type=HitasType.NEW_HITAS_I,
    )

    owner_2: Owner = OwnerFactory.create(name="Owner 2", non_disclosure=True)
    ownership_3: Ownership = OwnershipFactory.create(
        owner=owner_2,
        sale__apartment__building__real_estate__housing_company__postal_code__value="00001",
        sale__apartment__building__real_estate__housing_company__hitas_type=HitasType.NEW_HITAS_I,
    )
    ownership_4: Ownership = OwnershipFactory.create(
        owner=owner_2,
        sale__apartment__building__real_estate__housing_company__postal_code__value="00002",
        sale__apartment__building__real_estate__housing_company__hitas_type=HitasType.NEW_HITAS_I,
    )
    ownership_5: Ownership = OwnershipFactory.create(
        owner=owner_2,
        sale__apartment__building__real_estate__housing_company__postal_code__value="00003",
        sale__apartment__building__real_estate__housing_company__hitas_type=HitasType.NEW_HITAS_I,
    )

    url = reverse("hitas:multiple-ownerships-report-list")
    response: HttpResponse = api_client.get(url)

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert list(worksheet.values) == [
        (
            "Omistajan nimi",
            "Asunnon osoite",
            "Postinumero",
            "Omistajan asuntojen lukumäärä",
            "Omistajan henkilö- tai Y-tunnus",
            "Yhtiön nimi",
            "Yhtiön valmistumispäivä",
            "Kalleusalue",
        ),
        (
            ownership_1.owner.name,
            ownership_1.apartment.address,
            ownership_1.apartment.postal_code.value,
            2,
            None if ownership_1.owner.non_disclosure else ownership_1.owner.identifier,
            ownership_1.apartment.building.real_estate.housing_company.display_name,
            datetime.datetime.combine(
                ownership_1.apartment.building.real_estate.housing_company.completion_date, datetime.datetime.min.time()
            ),
            ownership_1.apartment.postal_code.cost_area,
        ),
        (
            ownership_2.owner.name,
            ownership_2.apartment.address,
            ownership_2.apartment.postal_code.value,
            2,
            None if ownership_2.owner.non_disclosure else ownership_2.owner.identifier,
            ownership_2.apartment.building.real_estate.housing_company.display_name,
            datetime.datetime.combine(
                ownership_2.apartment.building.real_estate.housing_company.completion_date, datetime.datetime.min.time()
            ),
            ownership_2.apartment.postal_code.cost_area,
        ),
        (
            Owner.OBFUSCATED_OWNER_NAME,
            ownership_3.apartment.address,
            ownership_3.apartment.postal_code.value,
            3,
            None if ownership_3.owner.non_disclosure else ownership_3.owner.identifier,
            ownership_3.apartment.building.real_estate.housing_company.display_name,
            datetime.datetime.combine(
                ownership_3.apartment.building.real_estate.housing_company.completion_date, datetime.datetime.min.time()
            ),
            ownership_3.apartment.postal_code.cost_area,
        ),
        (
            Owner.OBFUSCATED_OWNER_NAME,
            ownership_4.apartment.address,
            ownership_4.apartment.postal_code.value,
            3,
            None if ownership_4.owner.non_disclosure else ownership_4.owner.identifier,
            ownership_4.apartment.building.real_estate.housing_company.display_name,
            datetime.datetime.combine(
                ownership_4.apartment.building.real_estate.housing_company.completion_date, datetime.datetime.min.time()
            ),
            ownership_4.apartment.postal_code.cost_area,
        ),
        (
            Owner.OBFUSCATED_OWNER_NAME,
            ownership_5.apartment.address,
            ownership_5.apartment.postal_code.value,
            3,
            None if ownership_5.owner.non_disclosure else ownership_5.owner.identifier,
            ownership_5.apartment.building.real_estate.housing_company.display_name,
            datetime.datetime.combine(
                ownership_5.apartment.building.real_estate.housing_company.completion_date, datetime.datetime.min.time()
            ),
            ownership_5.apartment.postal_code.cost_area,
        ),
    ]


@pytest.mark.django_db
def test__api__regulated_ownerships_report__multiple_owners(api_client: HitasAPIClient):
    owner_1: Owner = OwnerFactory.create(name="Owner 1")
    OwnershipFactory.create(
        owner=owner_1,
        sale__apartment__building__real_estate__housing_company__postal_code__value="00001",
        sale__apartment__building__real_estate__housing_company__hitas_type=HitasType.NEW_HITAS_I,
    )
    OwnershipFactory.create(
        owner=owner_1,
        sale__apartment__building__real_estate__housing_company__postal_code__value="00002",
        sale__apartment__building__real_estate__housing_company__hitas_type=HitasType.NEW_HITAS_I,
    )

    owner_2: Owner = OwnerFactory.create(name="Owner 2", non_disclosure=True)
    OwnershipFactory.create(
        owner=owner_2,
        sale__apartment__building__real_estate__housing_company__postal_code__value="00001",
        sale__apartment__building__real_estate__housing_company__hitas_type=HitasType.NEW_HITAS_I,
    )
    OwnershipFactory.create(
        owner=owner_2,
        sale__apartment__building__real_estate__housing_company__postal_code__value="00002",
        sale__apartment__building__real_estate__housing_company__hitas_type=HitasType.NEW_HITAS_I,
    )
    OwnershipFactory.create(
        owner=owner_2,
        sale__apartment__building__real_estate__housing_company__postal_code__value="00003",
        sale__apartment__building__real_estate__housing_company__hitas_type=HitasType.NEW_HITAS_I,
    )

    url = reverse("hitas:regulated-ownerships-report-list")
    response: HttpResponse = api_client.get(url)

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]

    assert len(list(worksheet.values)) == 6, "There should be 5 ownership rows and 1 header row"


@pytest.mark.django_db
def test__api__download_ownerships_by_housing_company(api_client: HitasAPIClient):
    housing_company = HousingCompanyFactory(
        display_name="HCompany1",
    )
    ownership_1 = OwnershipFactory(
        sale__apartment__apartment_number=20,
        sale__apartment__surface_area=50.2,
        sale__apartment__building__real_estate__housing_company=housing_company,
    )
    non_disclosure_owner: Owner = OwnerFactory.create(name="Owner 2 non-disclosure", non_disclosure=True)
    OwnershipFactory(
        owner=non_disclosure_owner,
        sale__apartment=ownership_1.apartment,
        sale__purchase_date=datetime.date(2023, 10, 10),
    )

    ownership_2 = OwnershipFactory(
        sale__apartment__apartment_number=10,
        sale__apartment__surface_area=150.2,
        sale__apartment__share_number_start=10,
        sale__apartment__share_number_end=100,
        sale__apartment__building__real_estate__housing_company=housing_company,
        sale__purchase_date=datetime.date(2023, 9, 9),
    )

    url = reverse("hitas:download-ownership-by-housing-company-report-detail", kwargs={"pk": housing_company.uuid})
    response: HttpResponse = api_client.get(url)
    assert f"attachment; filename=Omistajat yhtiölle {housing_company.display_name}.xlsx" == response.headers.get(
        "Content-Disposition"
    )

    workbook: Workbook = load_workbook(BytesIO(response.content), data_only=False)
    worksheet: Worksheet = workbook.worksheets[0]
    assert list(worksheet.values) == [
        ("Asunnon nro", "Asunnon pinta-ala", "Osakenumerot", "Kauppakirjapäivä", "Omistajan nimi", "Henkilötunnus"),
        (
            10,
            150.2,
            "10-100",
            datetime.datetime.fromisoformat("2023-09-09"),
            ownership_2.owner.name,
            ownership_2.owner.identifier,
        ),
        (
            20,
            50.2,
            f"{ownership_1.sale.apartment.share_number_start}-{ownership_1.sale.apartment.share_number_end}",
            datetime.datetime.fromisoformat(ownership_1.sale.purchase_date.isoformat()),
            ownership_1.owner.name,
            ownership_1.owner.identifier,
        ),
        (
            20,
            50.2,
            f"{ownership_1.sale.apartment.share_number_start}-{ownership_1.sale.apartment.share_number_end}",
            datetime.datetime.fromisoformat("2023-10-10"),
            "***",
            " ",
        ),
    ]


@pytest.mark.parametrize("regulated", [False, True])
@pytest.mark.django_db
def test__api__ownerships_by_housing_company(api_client: HitasAPIClient, regulated):
    housing_company = HousingCompanyFactory(
        display_name="HCompany1",
        regulation_status=RegulationStatus.REGULATED if regulated else RegulationStatus.RELEASED_BY_HITAS,
    )
    ownership_1 = OwnershipFactory(
        sale__apartment__apartment_number=20,
        sale__apartment__surface_area=50.2,
        sale__apartment__building__real_estate__housing_company=housing_company,
    )
    non_disclosure_owner: Owner = OwnerFactory.create(name="Owner 2 non-disclosure", non_disclosure=True)
    ownership_1_2 = OwnershipFactory(
        owner=non_disclosure_owner,
        sale__apartment=ownership_1.apartment,
        sale__purchase_date=datetime.date(2023, 10, 10),
    )

    ownership_2 = OwnershipFactory(
        sale__apartment__apartment_number=10,
        sale__apartment__surface_area=150.2,
        sale__apartment__share_number_start=10,
        sale__apartment__share_number_end=100,
        sale__apartment__building__real_estate__housing_company=housing_company,
        sale__purchase_date=datetime.date(2023, 9, 9),
    )

    url = reverse("hitas:ownership-by-housing-company-report-detail", kwargs={"pk": housing_company.uuid})
    response = api_client.get(url)

    assert response.status_code == status.HTTP_200_OK

    if regulated:

        def get_share_range(apartment: Apartment):
            return f"{apartment.share_number_start}-{apartment.share_number_end}"

        assert response.json() == [
            {
                "owner_id": ownership_2.owner.uuid.hex,
                "number": 10,
                "owner_name": ownership_2.owner.name,
                "owner_ssn": ownership_2.owner.identifier,
                "purchase_date": "2023-09-09",
                "share_numbers": "10-100",
                "surface_area": 150.2,
            },
            {
                "owner_id": ownership_1.owner.uuid.hex,
                "number": 20,
                "owner_name": ownership_1.owner.name,
                "owner_ssn": ownership_1.owner.identifier,
                "purchase_date": str(ownership_1.sale.purchase_date),
                "share_numbers": get_share_range(ownership_1.sale.apartment),
                "surface_area": 50.2,
            },
            {
                "owner_id": ownership_1_2.owner.uuid.hex,
                "number": 20,
                "owner_name": "***",
                "owner_ssn": "",
                "purchase_date": "2023-10-10",
                "share_numbers": get_share_range(ownership_1_2.sale.apartment),
                "surface_area": 50.2,
            },
        ]
    else:
        assert response.json() == []
