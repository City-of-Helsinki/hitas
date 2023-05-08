import datetime
import json
import logging
from decimal import Decimal
from itertools import chain
from typing import TYPE_CHECKING, Iterable, Literal, NamedTuple, Optional, TypedDict
from uuid import UUID

from dateutil.relativedelta import relativedelta
from django.db import models
from django.db.models import Count, F, Max, Min, OuterRef, Prefetch, Q, Subquery
from django.db.models.functions import TruncMonth
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.worksheet.worksheet import Worksheet

from hitas.exceptions import HitasModelNotFound, ModelConflict, get_hitas_object_or_404
from hitas.models import HitasPostalCode
from hitas.models.apartment_sale import ApartmentSale
from hitas.models.external_sales_data import ExternalSalesData, SaleData
from hitas.models.housing_company import (
    HitasType,
    HousingCompany,
    HousingCompanyState,
    HousingCompanyWithAnnotations,
    RegulationStatus,
)
from hitas.models.indices import SurfaceAreaPriceCeiling
from hitas.models.thirty_year_regulation import (
    FullSalesData,
    HousingCompanyNameT,
    HousingCompanyUUIDHex,
    PostalCodeT,
    QuarterT,
    RegulationResult,
    ReplacementPostalCodesWithPrice,
    ThirtyYearRegulationResults,
    ThirtyYearRegulationResultsRow,
    ThirtyYearRegulationResultsRowWithAnnotations,
)
from hitas.services.housing_company import get_completed_housing_companies, make_index_adjustment_for_housing_companies
from hitas.services.indices import subquery_appropriate_cpi
from hitas.services.owner import obfuscate_owners_without_regulated_apartments
from hitas.utils import (
    business_quarter,
    format_sheet,
    hitas_calculation_quarter,
    humanize_relativedelta,
    resize_columns,
    roundup,
    subquery_count,
    to_quarter,
)

if TYPE_CHECKING:
    from hitas.models.owner import OwnerT

logger = logging.getLogger()


class AddressInfo(TypedDict):
    street_address: str
    postal_code: str
    city: str


class PropertyManagerInfo(TypedDict):
    id: str
    name: str
    email: str
    address: AddressInfo


class ComparisonData(TypedDict):
    id: HousingCompanyUUIDHex
    display_name: HousingCompanyNameT
    address: AddressInfo
    price: Decimal
    old_ruleset: bool
    completion_date: datetime.date | str
    property_manager: PropertyManagerInfo
    letter_fetched: bool
    current_regulation_status: RegulationStatus | str


class RegulationResults(TypedDict):
    automatically_released: list[ComparisonData]
    released_from_regulation: list[ComparisonData]
    stays_regulated: list[ComparisonData]
    skipped: list[ComparisonData]
    obfuscated_owners: list["OwnerT"]


class ReportColumns(NamedTuple):
    display_name: str
    acquisition_price: Decimal | str
    apartment_count: int | str
    indices: str
    change: Decimal | str
    adjusted_acquisition_price: Decimal | str
    surface_area: Decimal | str
    price_per_square_meter: Decimal | str
    postal_code_price: Decimal | str | None
    state: str
    completion_date: datetime.date | str
    age: str


class RegulationPostalCode(TypedDict):
    postal_code: str
    price_by_area: Decimal
    cost_area: Optional[int]


def perform_thirty_year_regulation(
    calculation_date: datetime.date,
    replacement_postal_codes: Optional[dict[PostalCodeT, list[PostalCodeT]]] = None,
) -> RegulationResults:
    """Check housing companies over 30 years old whether they should be released from hitas regulation or not.

    :param calculation_date: Date to check regulation from.
    :param replacement_postal_codes: Postal codes to replace with other postal codes in case some housing companies
                                     cannot be regulated due to missing sales data.
    """

    replacement_postal_codes = replacement_postal_codes or {}

    calculation_month = hitas_calculation_quarter(calculation_date)
    regulation_month = calculation_month - relativedelta(years=30)

    this_quarter = business_quarter(calculation_month)
    this_quarter_previous_year = this_quarter - relativedelta(years=1)
    previous_quarter = this_quarter - relativedelta(months=3)

    check_existing_regulation_data(calculation_month)

    logger.info(f"Checking regulation need for housing companies completed before {regulation_month.isoformat()!r}...")

    logger.info("Fetching housing companies...")
    housing_companies = get_completed_housing_companies(completion_month=regulation_month)
    if not housing_companies:
        logger.info("No housing companies to check regulation for.")
        logger.info("Regulation check complete!")
        return RegulationResults(
            automatically_released=[],
            released_from_regulation=[],
            stays_regulated=[],
            skipped=[],
            obfuscated_owners=[],
        )

    split_housing_companies, automatically_released = _split_automatically_released(housing_companies)
    if not housing_companies:
        logger.info("All housing companies qualify for automatic release.")
        results = RegulationResults(
            automatically_released=automatically_released,
            released_from_regulation=[],
            stays_regulated=[],
            skipped=[],
            obfuscated_owners=[],
        )

        logger.info("Updating housing company states...")
        _free_housing_companies_from_regulation(split_housing_companies, results)

        results["obfuscated_owners"] = obfuscate_owners_without_regulated_apartments()

        logger.info("Saving regulation results for reporting...")
        _save_regulation_results(
            results=results,
            calculation_month=calculation_month,
            regulation_month=regulation_month,
            housing_companies=split_housing_companies,
            surface_area_price_ceiling=None,
            unadjusted_prices=None,
            indices=None,
            sales_data=None,
            external_sales_data=None,
            price_by_area=None,
            replacement_postal_codes=replacement_postal_codes,
        )

        logger.info("Regulation check complete!")
        return results

    logger.info(
        f"{len(split_housing_companies)} housing companies qualify for automatic release. "
        f"Proceeding with regulation checks for remaining {len(housing_companies)} housing companies..."
    )

    unadjusted_prices: dict[HousingCompanyUUIDHex, Decimal] = {
        housing_company.uuid.hex: housing_company.avg_price_per_square_meter for housing_company in housing_companies
    }

    logger.info("Making index adjustments...")
    indices = make_index_adjustment_for_housing_companies(housing_companies, calculation_month)

    logger.info(f"Fetching surface area price ceiling for {calculation_month.isoformat()!r}...")
    surface_area_price_ceiling = get_hitas_object_or_404(SurfaceAreaPriceCeiling, month=calculation_month)

    logger.info("Determining comparison values for housing companies...")
    comparison_values = _get_comparison_values(housing_companies, surface_area_price_ceiling.value)

    postal_codes: set[PostalCodeT] = set(comparison_values) | {
        postal_code for postal_codes in replacement_postal_codes.values() for postal_code in postal_codes
    }

    logger.info(
        f"Fetching HITAS sales data between {this_quarter_previous_year.isoformat()} (inclusive) "
        f"and {this_quarter.isoformat()} (exclusive)..."
    )
    sales_data = get_sales_data(this_quarter_previous_year, this_quarter, postal_codes)

    logger.info("Fetching external sales data for the last four quarters...")
    external_sales_data = get_external_sales_data(to_quarter(previous_quarter), postal_codes)

    logger.info("Combining HITAS and external sales data for each postal code...")
    price_by_area = combine_sales_data(sales_data, external_sales_data)

    logger.info("Determining regulation need for housing companies...")
    results = _determine_regulation_need(comparison_values, price_by_area, replacement_postal_codes)

    if results["skipped"]:
        logger.info(
            f"{len(results['skipped'])} housing companies could not be checked. Regulation could not be completed."
        )
        return RegulationResults(
            automatically_released=[],
            released_from_regulation=[],
            stays_regulated=[],
            skipped=results["skipped"],
            obfuscated_owners=[],
        )

    results["automatically_released"] += automatically_released

    logger.info(
        f"{len(results['automatically_released'])} housing companies are released from regulation automatically, "
        f"{len(results['released_from_regulation'])} are released after regulation checks, "
        f"{len(results['stays_regulated'])} stay regulated."
    )

    housing_companies += split_housing_companies

    logger.info("Updating housing company states...")
    _free_housing_companies_from_regulation(housing_companies, results)

    results["obfuscated_owners"] = obfuscate_owners_without_regulated_apartments()

    logger.info("Saving regulation results for reporting...")
    _save_regulation_results(
        results=results,
        calculation_month=calculation_month,
        regulation_month=regulation_month,
        housing_companies=housing_companies,
        surface_area_price_ceiling=surface_area_price_ceiling.value,
        unadjusted_prices=unadjusted_prices,
        indices=indices,
        sales_data=sales_data,
        external_sales_data=external_sales_data,
        price_by_area=price_by_area,
        replacement_postal_codes=replacement_postal_codes,
    )

    logger.info("Regulation check complete!")
    return results


def check_existing_regulation_data(calculation_month: datetime.date) -> None:
    """
    Check if there is already regulation results for this hitas quarter.
    Do not allow re-regulation for a given quarter.
    """
    if ThirtyYearRegulationResults.objects.filter(calculation_month=calculation_month).exists():
        raise ModelConflict(
            "Previous regulation exists. Cannot re-check regulation for this quarter.",
            error_code="unique",
        )


def _split_automatically_released(
    housing_companies: list[HousingCompanyWithAnnotations],
) -> tuple[list[HousingCompanyWithAnnotations], list[ComparisonData]]:
    """
    Separate out housing companies that are automatically released from regulation.
    """
    split_housing_companies: list[HousingCompanyWithAnnotations] = []
    automatically_released: list[ComparisonData] = []
    for i, housing_company in enumerate(housing_companies):
        if not housing_company.hitas_type.old_hitas_ruleset:
            logger.info(
                f"Housing company {housing_company.display_name!r} uses new hitas ruleset. "
                f"Automatically qualified for release from regulation."
            )
            split_housing_companies.append(housing_company)
            automatically_released.append(
                ComparisonData(
                    id=housing_company.uuid.hex,
                    display_name=housing_company.display_name,
                    address=AddressInfo(
                        street_address=housing_company.street_address,
                        postal_code=housing_company.postal_code.value,
                        city=housing_company.postal_code.city,
                    ),
                    price=Decimal("0"),  # Index adjusted price not calculate yet
                    old_ruleset=housing_company.hitas_type.old_hitas_ruleset,
                    completion_date=housing_company.completion_date,
                    property_manager=PropertyManagerInfo(
                        id=housing_company.property_manager.uuid.hex,
                        name=housing_company.property_manager.name,
                        email=housing_company.property_manager.email,
                        address=AddressInfo(
                            street_address=housing_company.property_manager.street_address,
                            postal_code=housing_company.property_manager.postal_code,
                            city=housing_company.property_manager.city,
                        ),
                    ),
                    letter_fetched=False,
                    current_regulation_status=RegulationStatus.RELEASED_BY_HITAS.value,
                )
            )
            housing_companies[i] = None

    housing_companies[:] = [housing_company for housing_company in housing_companies if housing_company is not None]
    return split_housing_companies, automatically_released


def _get_comparison_values(
    housing_companies: list[HousingCompanyWithAnnotations],
    surface_area_price_ceiling: Decimal,
) -> dict[PostalCodeT, dict[HousingCompanyUUIDHex, ComparisonData]]:
    """
    Determine the value, for each housing company, which shall be compared against
    the housing company's postal area's final average price per square meter.
    """
    comparison_values: dict[PostalCodeT, dict[HousingCompanyUUIDHex, ComparisonData]] = {}
    for housing_company in housing_companies:
        postal_code = housing_company.postal_code.value
        comparison_values.setdefault(postal_code, {})
        comparison_values[postal_code][housing_company.uuid.hex] = ComparisonData(
            id=housing_company.uuid.hex,
            display_name=housing_company.display_name,
            address=AddressInfo(
                street_address=housing_company.street_address,
                postal_code=housing_company.postal_code.value,
                city=housing_company.postal_code.city,
            ),
            price=max((housing_company.avg_price_per_square_meter, surface_area_price_ceiling)),
            old_ruleset=housing_company.hitas_type.old_hitas_ruleset,
            completion_date=housing_company.completion_date,
            property_manager=PropertyManagerInfo(
                id=housing_company.property_manager.uuid.hex,
                name=housing_company.property_manager.name,
                email=housing_company.property_manager.email,
                address=AddressInfo(
                    street_address=housing_company.property_manager.street_address,
                    postal_code=housing_company.property_manager.postal_code,
                    city=housing_company.property_manager.city,
                ),
            ),
            letter_fetched=False,
            current_regulation_status=RegulationStatus.REGULATED.value,  # Changed later if necessary
        )

    return comparison_values


def get_sales_data(
    from_: datetime.date,
    to_: datetime.date,
    postal_codes: Optional[set[PostalCodeT]] = None,
) -> dict[PostalCodeT, dict[QuarterT, SaleData]]:
    """
    Find all sales for apartments (in the given postal codes if not None) between the given dates
    ('from_' is inclusive, 'to_' is exclusive) for use in regulation check.
    """

    sales_by_quarter: dict[PostalCodeT, dict[QuarterT, SaleData]] = {}
    sales_in_previous_year = (
        ApartmentSale.objects.select_related(
            "apartment",
            "apartment__building",
            "apartment__building__real_estate",
            "apartment__building__real_estate__housing_company",
            "apartment__building__real_estate__housing_company__postal_code",
        )
        .alias(
            _first_sale_id=Subquery(
                queryset=(
                    ApartmentSale.objects.filter(apartment_id=OuterRef("apartment_id"))
                    .order_by("purchase_date", "id")
                    .values_list("id", flat=True)[:1]
                ),
                output_field=models.IntegerField(null=True),
            ),
        )
        .filter(
            ~Q(id__in=F("_first_sale_id")),
            ~Q(apartment__building__real_estate__housing_company__hitas_type=HitasType.HALF_HITAS),
            purchase_date__gte=from_,
            purchase_date__lt=to_,
            exclude_from_statistics=False,
            apartment__building__real_estate__housing_company__exclude_from_statistics=False,
        )
    )
    if postal_codes is not None:
        sales_in_previous_year = sales_in_previous_year.filter(
            apartment__building__real_estate__housing_company__postal_code__value__in=postal_codes,
        )

    for sale in sales_in_previous_year:
        quarter = to_quarter(sale.purchase_date)
        postal_code = sale.apartment.postal_code.value
        if postal_codes is not None and postal_code not in postal_codes:
            continue

        sales_by_quarter.setdefault(postal_code, {})
        sales_by_quarter[postal_code].setdefault(quarter, SaleData(sale_count=0, price=0))
        sales_by_quarter[postal_code][quarter]["sale_count"] += 1
        sales_by_quarter[postal_code][quarter]["price"] += sale.total_price

    return sales_by_quarter


def get_external_sales_data(
    quarter: QuarterT,
    postal_codes: Optional[set[PostalCodeT]] = None,
) -> dict[PostalCodeT, dict[QuarterT, SaleData]]:
    """
    Get all external sales data (in the given postal codes if not None) for the given quarter.
    """
    sales_data = get_hitas_object_or_404(ExternalSalesData, calculation_quarter=quarter)

    sales_by_quarter: dict[PostalCodeT, dict[QuarterT, SaleData]] = {}
    for quarter_data in (sales_data.quarter_1, sales_data.quarter_2, sales_data.quarter_3, sales_data.quarter_4):
        for area in quarter_data["areas"]:
            quarter_ = quarter_data["quarter"]
            postal_code = area["postal_code"]
            if postal_codes is not None and postal_code not in postal_codes:
                continue

            sales_by_quarter.setdefault(postal_code, {})
            sales_by_quarter[postal_code][quarter_] = SaleData(sale_count=area["sale_count"], price=area["price"])

    return sales_by_quarter


def combine_sales_data(*args: dict[PostalCodeT, dict[QuarterT, SaleData]]) -> dict[PostalCodeT, Decimal]:
    """
    Combine multiple sources of sales data (e.g. Hitas sales and External sales data),
    and calculate the average per postal code.
    """
    sales_data: dict[PostalCodeT, SaleData] = {}
    for data in args:
        for postal_code, quarter_data in data.items():
            for sale_data in quarter_data.values():
                sales_data.setdefault(postal_code, SaleData(sale_count=0, price=0))
                sales_data[postal_code]["sale_count"] += sale_data["sale_count"]
                sales_data[postal_code]["price"] += sale_data["price"]

    total_by_postal_code: dict[PostalCodeT, Decimal] = {}
    for postal_code, sale_data in sales_data.items():
        total_by_postal_code[postal_code] = roundup(Decimal(sale_data["price"]) / Decimal(sale_data["sale_count"]))

    return total_by_postal_code


def _determine_regulation_need(
    comparison_values: dict[PostalCodeT, dict[HousingCompanyUUIDHex, ComparisonData]],
    price_by_area: dict[PostalCodeT, Decimal],
    replacement_postal_codes: dict[PostalCodeT, list[PostalCodeT]],
) -> RegulationResults:
    """
    Determine id a given housing company should stay regulated or be released from regulation
    based on its "comparison value" (determined separately).
    """
    results = RegulationResults(
        automatically_released=[],
        released_from_regulation=[],
        stays_regulated=[],
        skipped=[],
        obfuscated_owners=[],
    )

    if replacement_postal_codes:
        missing: list[PostalCodeT] = [
            postal_code
            for postal_codes in replacement_postal_codes.values()
            for postal_code in postal_codes
            if postal_code not in price_by_area
        ]
        if missing:
            raise ModelConflict(
                f"Missing price data for replacement postal codes: {missing}.",
                error_code="missing",
            )

    for postal_code, comparison_data_by_housing_company in comparison_values.items():
        postal_code_average_price_per_square_meter = price_by_area.get(postal_code)

        if postal_code_average_price_per_square_meter is None:
            postal_code_average_price_per_square_meter = _find_average_of_nearest(
                postal_code,
                price_by_area,
                replacement_postal_codes,
            )
            if postal_code_average_price_per_square_meter is not None:
                logger.info(
                    f"Using average price per square meter from replacement postal codes: "
                    f"{postal_code} -> {replacement_postal_codes[postal_code]}."
                )

        for comparison_data in comparison_data_by_housing_company.values():
            display_name = comparison_data["display_name"]
            comparison_value = comparison_data["price"]

            if postal_code_average_price_per_square_meter is None:
                logger.info(
                    f"No average price per square meter found for postal code {postal_code!r}, "
                    f"cannot determine regulation need for {display_name!r}."
                )
                results["skipped"].append(comparison_data)
                continue

            if comparison_value >= postal_code_average_price_per_square_meter:
                logger.info(
                    f"Housing company {display_name!r} should be released from regulation since: "
                    f"{comparison_value} >= {postal_code_average_price_per_square_meter}."
                )
                comparison_data["current_regulation_status"] = RegulationStatus.RELEASED_BY_HITAS.value
                results["released_from_regulation"].append(comparison_data)
            else:
                logger.info(
                    f"Housing company {display_name!r} should stay regulated since: "
                    f"{comparison_value} < {postal_code_average_price_per_square_meter}."
                )
                results["stays_regulated"].append(comparison_data)

    return results


def _find_average_of_nearest(
    postal_code: PostalCodeT,
    price_by_area: dict[PostalCodeT, Decimal],
    replacement_postal_codes: dict[PostalCodeT, list[PostalCodeT]],
) -> Optional[Decimal]:
    """
    Find the average of the nearest two postal areas in the given prices by area.
    """
    replacements = replacement_postal_codes.get(postal_code)
    if replacements is None:
        return None

    return sum(price_by_area[postal_code] for postal_code in replacements) / len(replacements)


def _free_housing_companies_from_regulation(
    housing_companies: list[HousingCompanyWithAnnotations],
    regulation_results: RegulationResults,
) -> None:
    """
    Change housing company states according to regulation results.
    """
    housing_company_uuids: set[str] = {
        results["id"]
        for results in chain(
            regulation_results["automatically_released"], regulation_results["released_from_regulation"]
        )
    }
    for housing_company in housing_companies:
        if housing_company.uuid.hex not in housing_company_uuids:
            housing_company.regulation_status = RegulationStatus.REGULATED
            housing_company.state = HousingCompanyState.GREATER_THAN_30_YEARS_NOT_FREE
            continue

        housing_company.regulation_status = RegulationStatus.RELEASED_BY_HITAS
        housing_company.state = HousingCompanyState.GREATER_THAN_30_YEARS_FREE

    HousingCompany.objects.bulk_update(housing_companies, fields=["state", "regulation_status"])


def _save_regulation_results(
    results: RegulationResults,
    calculation_month: datetime.date,
    regulation_month: datetime.date,
    housing_companies: Iterable[HousingCompanyWithAnnotations],
    surface_area_price_ceiling: Optional[Decimal],
    unadjusted_prices: Optional[dict[HousingCompanyUUIDHex, Decimal]],
    indices: Optional[dict[Literal["old", "new"], dict[datetime.date, Decimal]]],
    sales_data: Optional[dict[PostalCodeT, dict[QuarterT, SaleData]]],
    external_sales_data: Optional[dict[PostalCodeT, dict[QuarterT, SaleData]]],
    price_by_area: Optional[dict[PostalCodeT, Decimal]],
    replacement_postal_codes: dict[PostalCodeT, list[PostalCodeT]],
) -> ThirtyYearRegulationResults:
    """
    Save regulation results for reporting.
    """
    thirty_year_regulation_results, created = ThirtyYearRegulationResults.objects.update_or_create(
        calculation_month=calculation_month,
        defaults={
            "regulation_month": regulation_month,
            "surface_area_price_ceiling": surface_area_price_ceiling,
            "sales_data": FullSalesData(
                # Convert decimals to floats
                internal=json.loads(json.dumps(sales_data, default=float)) or {},
                external=json.loads(json.dumps(external_sales_data, default=float)) or {},
                price_by_area=json.loads(json.dumps(price_by_area, default=float)) or {},
            ),
            "replacement_postal_codes": [
                ReplacementPostalCodesWithPrice(
                    postal_code=postal_code,
                    replacements=replacements,
                    price_by_area=float(_find_average_of_nearest(postal_code, price_by_area, replacement_postal_codes)),
                )
                for postal_code, replacements in replacement_postal_codes.items()
            ],
        },
    )

    # Regulations are immutable, so they shouldn't be overridden.
    if not created:
        raise ModelConflict("Regulation results already exist for this month.", error_code="unique")

    rows_to_save: list[ThirtyYearRegulationResultsRow] = []
    for housing_company in housing_companies:
        # If housing company was released automatically, its prices were never index adjusted.
        unadjusted_average_price_per_square_meter: Decimal = housing_company.avg_price_per_square_meter
        adjusted_average_price_per_square_meter: Optional[Decimal] = None
        if unadjusted_prices is not None and housing_company.uuid.hex in unadjusted_prices:
            unadjusted_average_price_per_square_meter = unadjusted_prices[housing_company.uuid.hex]
            adjusted_average_price_per_square_meter = housing_company.avg_price_per_square_meter

        # If housing company was released automatically, its prices were never index adjusted.
        completion_month_index: Optional[Decimal] = None
        calculation_month_index: Optional[Decimal] = None
        key: Literal["old", "new"]
        key = "old" if housing_company.hitas_type.old_hitas_ruleset else "new"  # type: ignore
        if indices is not None and housing_company.completion_month in indices[key]:
            completion_month_index = indices[key][housing_company.completion_month]
            calculation_month_index = indices[key][calculation_month]

        if housing_company.uuid.hex in {result["id"] for result in results["automatically_released"]}:
            regulation_result = RegulationResult.AUTOMATICALLY_RELEASED
        elif housing_company.uuid.hex in {result["id"] for result in results["released_from_regulation"]}:
            regulation_result = RegulationResult.RELEASED_FROM_REGULATION
        else:
            regulation_result = RegulationResult.STAYS_REGULATED

        rows_to_save.append(
            ThirtyYearRegulationResultsRow(
                parent=thirty_year_regulation_results,
                housing_company=housing_company,
                completion_date=housing_company.completion_date,
                surface_area=housing_company.surface_area,
                postal_code=housing_company.postal_code.value,
                realized_acquisition_price=housing_company.realized_acquisition_price,
                unadjusted_average_price_per_square_meter=unadjusted_average_price_per_square_meter,
                adjusted_average_price_per_square_meter=adjusted_average_price_per_square_meter,
                completion_month_index=completion_month_index,
                calculation_month_index=calculation_month_index,
                regulation_result=regulation_result,
            )
        )

    ThirtyYearRegulationResultsRow.objects.bulk_create(objs=rows_to_save)

    return thirty_year_regulation_results


def get_thirty_year_regulation_results(calculation_date: datetime.date) -> ThirtyYearRegulationResults:
    try:
        return ThirtyYearRegulationResults.objects.prefetch_related(
            Prefetch(
                "rows",
                ThirtyYearRegulationResultsRow.objects.prefetch_related(
                    "housing_company",
                    "housing_company__postal_code",
                    "housing_company__property_manager",
                )
                .annotate(
                    apartment_count=Count("housing_company__real_estates__buildings__apartments"),
                )
                .order_by("regulation_result", "completion_date"),
            ),
        ).get(calculation_month=hitas_calculation_quarter(calculation_date))
    except ThirtyYearRegulationResults.DoesNotExist as error:
        raise HitasModelNotFound(ThirtyYearRegulationResults) from error


def get_thirty_year_regulation_results_for_housing_company(
    housing_company_uuid: UUID,
    calculation_date: datetime.date,
) -> ThirtyYearRegulationResultsRowWithAnnotations:
    results = (
        ThirtyYearRegulationResultsRow.objects.select_related(
            "parent",
            "housing_company",
            "housing_company__postal_code",
        )
        .prefetch_related(
            "housing_company__real_estates",
        )
        .filter(
            housing_company__uuid=housing_company_uuid,
            parent__calculation_month=hitas_calculation_quarter(calculation_date),
        )
        .annotate(
            check_count=subquery_count(
                ThirtyYearRegulationResultsRow,
                "housing_company__uuid",
                parent__calculation_month__lte=hitas_calculation_quarter(calculation_date),
            ),
            min_share=Min("housing_company__real_estates__buildings__apartments__share_number_start"),
            max_share=Max("housing_company__real_estates__buildings__apartments__share_number_end"),
            share_count=F("max_share") - F("min_share") + 1,
            apartment_count=Count("housing_company__real_estates__buildings__apartments"),
            completion_month=TruncMonth("completion_date"),
            completion_month_index_cpi=subquery_appropriate_cpi(
                outer_ref="completion_month",
                housing_company_ref="housing_company",
            ),
            calculation_month_index_cpi=subquery_appropriate_cpi(
                outer_ref="parent__calculation_month",
                housing_company_ref="housing_company",
            ),
            average_price_per_square_meter_cpi=(
                F("calculation_month_index_cpi") / F("completion_month_index_cpi") * F("realized_acquisition_price")
            ),
        )
        .order_by("-parent__calculation_month")
        .first()
    )
    if results is None:
        raise HitasModelNotFound(ThirtyYearRegulationResultsRow)

    results.turned_30 = results.completion_date + relativedelta(years=30)
    results.difference = results.adjusted_average_price_per_square_meter - Decimal(
        results.parent.sales_data["price_by_area"].get(results.postal_code, 0)
    )

    return results


def build_thirty_year_regulation_report_excel(results: ThirtyYearRegulationResults) -> Workbook:
    workbook = Workbook()
    worksheet: Worksheet = workbook.active

    columns = ReportColumns(
        display_name="Yhtiö",
        acquisition_price="Hankinta-arvo",
        apartment_count="Huoneistoja",
        indices="Indeksit",
        change="Muutos",
        adjusted_acquisition_price="Takistettu hinta",
        surface_area="Pinta-ala",
        price_per_square_meter="E-hinta/m²",
        postal_code_price="Postinumerohinta",
        state="Tila",
        completion_date="Valmistumispäivä",
        age="Yhtiön ikä",
    )
    worksheet.append(columns)

    for row in results.rows.all():
        data = ReportColumns(
            display_name=row.housing_company.display_name,
            acquisition_price=row.realized_acquisition_price,
            apartment_count=row.apartment_count,
            indices=f"{row.completion_month_index}/{row.calculation_month_index}",
            change=(
                (row.adjusted_average_price_per_square_meter * row.surface_area)
                - (row.unadjusted_average_price_per_square_meter * row.surface_area)
            ),
            adjusted_acquisition_price=row.adjusted_average_price_per_square_meter * row.surface_area,
            surface_area=row.surface_area,
            price_per_square_meter=row.adjusted_average_price_per_square_meter,
            postal_code_price=row.parent.sales_data["price_by_area"].get(row.postal_code),
            state=("Ei vapaudu" if row.regulation_result == RegulationResult.STAYS_REGULATED else "Vapautuu"),
            completion_date=row.completion_date,
            age=humanize_relativedelta(relativedelta(results.calculation_month, row.completion_date)),
        )
        worksheet.append(data)

    last_row = worksheet.max_row
    worksheet.auto_filter.ref = worksheet.dimensions

    # There needs to be an empty row for sorting and filtering to work properly
    worksheet.append(
        ReportColumns(
            display_name="",
            acquisition_price="",
            apartment_count="",
            indices="",
            change="",
            adjusted_acquisition_price="",
            surface_area="",
            price_per_square_meter="",
            postal_code_price="",
            state="",
            completion_date="",
            age="",
        )
    )

    summary_start = worksheet.max_row + 1
    summary_rows = {"Summa": "SUM", "Keskiarvo": "AVERAGE", "Mediaani": "MEDIAN"}
    for title, formula in summary_rows.items():
        worksheet.append(
            ReportColumns(
                display_name=title,
                acquisition_price=f"={formula}(B2:B{last_row})",
                apartment_count=f"={formula}(C2:C{last_row})",
                indices="",
                change=f"={formula}(E2:E{last_row})",
                adjusted_acquisition_price=f"={formula}(F2:F{last_row})",
                surface_area=f"={formula}(G2:G{last_row})",
                price_per_square_meter=f"={formula}(H2:H{last_row})",
                postal_code_price=f"={formula}(I2:I{last_row})",
                state="",
                completion_date="",
                age="",
            )
        )

    format_sheet(
        worksheet,
        formatting_rules={
            # Add a border to the header row
            **{f"{letter}1": {"border": Border(bottom=Side(style="thin"))} for letter in "ABCDEFGHIJKL"},
            # Add a border to the last data row
            **{f"{letter}{last_row}": {"border": Border(bottom=Side(style="thin"))} for letter in "ABCDEFGHIJKL"},
            # Align the summary titles to the right
            **{
                f"A{summary_start + i}": {"alignment": Alignment(horizontal="right")}
                for i in range(0, len(summary_rows))
            },
            "B": {"number_format": "#,##0.00\\ €"},
            "D": {"alignment": Alignment(horizontal="right")},
            "E": {"number_format": "#,##0.00\\ €"},
            "F": {"number_format": "#,##0.00\\ €"},
            "G": {"number_format": "#,##0.00\\ \\m\\²"},
            "H": {"number_format": "#,##0.00\\ €"},
            "I": {"number_format": "#,##0.00\\ €"},
            "J": {
                "alignment": Alignment(horizontal="right"),
                # Change the font color depending on the text in the field
                "font": {
                    "Ei vapaudu": Font(color="FF0000"),  # red
                    "Vapautuu": Font(color="00FF00"),  # green
                },
            },
            "K": {"number_format": "DD.MM.YYYY"},
            "L": {"alignment": Alignment(horizontal="right")},
        },
    )

    resize_columns(worksheet)
    worksheet.protection.sheet = True
    return workbook


def convert_thirty_year_regulation_results_to_comparison_data(
    results: ThirtyYearRegulationResults,
) -> RegulationResults:
    regulation_results = RegulationResults(
        automatically_released=[],
        released_from_regulation=[],
        stays_regulated=[],
        skipped=[],
        obfuscated_owners=[],
    )
    for row in results.rows.all():
        column: dict[
            RegulationResult,
            Literal["automatically_released", "released_from_regulation", "stays_regulated"],
        ]
        column = {
            RegulationResult.AUTOMATICALLY_RELEASED: "automatically_released",
            RegulationResult.RELEASED_FROM_REGULATION: "released_from_regulation",
            RegulationResult.STAYS_REGULATED: "stays_regulated",
        }

        regulation_results[column[row.regulation_result]].append(
            ComparisonData(
                id=row.housing_company.uuid.hex,
                display_name=row.housing_company.display_name,
                address=AddressInfo(
                    street_address=row.housing_company.street_address,
                    postal_code=row.housing_company.postal_code.value,
                    city=row.housing_company.postal_code.city,
                ),
                price=max(
                    (row.adjusted_average_price_per_square_meter or 0), (results.surface_area_price_ceiling or 0)
                ),
                old_ruleset=row.housing_company.hitas_type.old_hitas_ruleset,
                completion_date=row.completion_date,
                property_manager=PropertyManagerInfo(
                    id=row.housing_company.property_manager.uuid.hex,
                    name=row.housing_company.property_manager.name,
                    email=row.housing_company.property_manager.email,
                    address=AddressInfo(
                        street_address=row.housing_company.property_manager.street_address,
                        postal_code=row.housing_company.property_manager.postal_code,
                        city=row.housing_company.property_manager.city,
                    ),
                ),
                letter_fetched=row.letter_fetched,
                current_regulation_status=row.housing_company.regulation_status.value,
            )
        )

    return regulation_results


def compile_postal_codes_with_cost_areas(price_by_area: dict[PostalCodeT, Decimal]) -> list[RegulationPostalCode]:
    results: dict[PostalCodeT, RegulationPostalCode] = {
        postal_code: RegulationPostalCode(
            postal_code=postal_code,
            price_by_area=price_by_area,
            cost_area=None,
        )
        for postal_code, price_by_area in price_by_area.items()
    }

    for postal_code in HitasPostalCode.objects.filter(value__in=list(price_by_area)):
        results[postal_code.value]["cost_area"] = postal_code.cost_area

    return list(results.values())
