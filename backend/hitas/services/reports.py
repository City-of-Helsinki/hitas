import datetime
import string
from decimal import Decimal
from enum import Enum
from functools import cache
from statistics import mean
from typing import Any, Callable, Iterable, Literal, NamedTuple, TypeAlias, TypedDict, TypeVar, Union

from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from rest_framework.exceptions import ValidationError
from rest_framework.settings import api_settings

from hitas.models import ApartmentSale, Owner
from hitas.models.housing_company import (
    HitasType,
    HousingCompany,
    HousingCompanyWithRegulatedReportAnnotations,
    HousingCompanyWithStateReportAnnotations,
    HousingCompanyWithUnregulatedReportAnnotations,
    RegulationStatus,
)
from hitas.models.indices import SurfaceAreaPriceCeiling
from hitas.models.ownership import Ownership, OwnershipWithApartmentCount
from hitas.models.property_manager import PropertyManager
from hitas.utils import format_sheet, resize_columns

T = TypeVar("T")
CostAreaT: TypeAlias = Literal[1, 2, 3, 4]
PostalCodeT: TypeAlias = str  # e.g. '00100'
RoomLabelT: TypeAlias = Literal["1h", "2h", "3h+"]


class SalesReportColumns(NamedTuple):
    cost_area: int | str
    postal_code: str
    apartment_address: str
    notification_date: datetime.date | str
    purchase_date: datetime.date | str
    purchase_price_per_square_meter: Decimal | str
    total_price_per_square_meter: Decimal | str


class SalesAndMaximumPricesReportColumns(NamedTuple):
    cost_area: int | str
    postal_code: str
    apartment_address: str
    surface_area_square_meter: Decimal | str
    purchase_date: datetime.date | str
    total_price: Decimal | str
    total_price_per_square_meter: Decimal | str
    maximum_price: Decimal | str
    maximum_price_per_square_meter: Decimal | str


class RegulatedHousingCompaniesReportColumns(NamedTuple):
    cost_area: int | str
    postal_code: str
    housing_company_name: str
    address: str
    completion_date: datetime.date | str
    apartment_count: int | str
    average_price_per_square_meter: Decimal | str


class UnregulatedHousingCompaniesReportColumns(NamedTuple):
    housing_company_name: str
    postal_code: str | int
    completion_date: datetime.date | str
    release_date: datetime.date | str
    released_by_ploy_department: str
    apartment_count: int | str


class PropertyManagerReportColumns(NamedTuple):
    property_manager_name: str
    property_manager_email: str
    housing_company_name: str


class HousingCompanyStatesCount(TypedDict):
    housing_company_count: int
    apartment_count: int


class HousingCompanyStatesReportColumns(NamedTuple):
    state: str
    housing_company_count: int | str
    apartment_count: int | str


class SalesReportSummaryDefinition(NamedTuple):
    func: Callable[[Iterable[T]], T]
    title: str = ""
    subtitle: str = ""


class ReportState(str, Enum):
    NOT_READY = "Ei valmis"
    REGULATED = "Sääntelyn piirissä"
    RELEASED_BY_HITAS = "Sääntelystä vapautuneet"
    RELEASED_BY_PLOT_DEPARTMENT = "Vapautuneet tontit-yksikön päätöksellä"
    HALF_HITAS = "Puolihitas yhtiöt"


class SalesInfo(TypedDict):
    sales_count: int
    sum: Decimal
    minimum: Decimal
    maximum: Decimal


class OwnershipReportColumns(NamedTuple):
    owner_name: str
    apartment_address: str
    postal_code: str
    owner_identifier: str
    housing_company_name: str
    housing_company_completion_date: datetime.date | str
    cost_area: int | str


class MultipleOwnershipReportColumns(NamedTuple):
    owner_name: str
    apartment_address: str
    postal_code: str
    apartment_count: int | str
    owner_identifier: str
    housing_company_name: str
    housing_company_completion_date: datetime.date | str
    cost_area: int | str


class OwnersByHousingCompanyReportColumns(NamedTuple):
    number: str | int
    surface_area: str | Decimal
    share_numbers: str
    purchase_date: str
    owner_name: str
    owner_ssn: str


SalesInfoByRoomCount = TypedDict(
    "SalesInfoByRoomCount",
    {
        "1h": SalesInfo,
        "2h": SalesInfo,
        "3h+": SalesInfo,
    },
    total=False,
)


class SalesByCostAreaColumns(NamedTuple):
    cost_area: CostAreaT | str
    postal_code: PostalCodeT | str
    room_label: RoomLabelT | str
    sales_count: int | str
    average_price_per_square_meter: Decimal | str
    minimum_price_per_square_meter: Decimal | str
    maximum_price_per_square_meter: Decimal | str


class SalesByCostArea(NamedTuple):
    sales_by_cost_area: dict[CostAreaT, dict[PostalCodeT, SalesInfoByRoomCount]]
    overall_count: int
    overall_average: Decimal
    overall_minimum: Decimal
    overall_maximum: Decimal


def build_sales_report_excel(sales: list[ApartmentSale]) -> Workbook:
    workbook = Workbook()
    worksheet: Worksheet = workbook.active

    column_headers = SalesReportColumns(
        cost_area="Kalleusalue",
        postal_code="Postinumero",
        apartment_address="Osoite",
        notification_date="Ilmoituspäivä",
        purchase_date="Kauppapäivä",
        purchase_price_per_square_meter="Kaupan neliöhinta",
        total_price_per_square_meter="Velaton neliöhinta",
    )
    worksheet.append(column_headers)

    for sale in sales:
        if not sale.apartment.surface_area:
            raise ValidationError(
                detail={
                    api_settings.NON_FIELD_ERRORS_KEY: (
                        f"Surface area zero or missing for apartment {sale.apartment.address!r}. "
                        f"Cannot calculate price per square meter."
                    )
                },
            )

        worksheet.append(
            SalesReportColumns(
                cost_area=sale.apartment.postal_code.cost_area,
                postal_code=sale.apartment.postal_code.value,
                apartment_address=sale.apartment.address,
                notification_date=sale.notification_date,
                purchase_date=sale.purchase_date,
                purchase_price_per_square_meter=sale.purchase_price / sale.apartment.surface_area,
                total_price_per_square_meter=sale.total_price / sale.apartment.surface_area,
            )
        )

    last_row = worksheet.max_row
    worksheet.auto_filter.ref = worksheet.dimensions

    empty_row = SalesReportColumns(
        cost_area="",
        postal_code="",
        apartment_address="",
        notification_date="",
        purchase_date="",
        purchase_price_per_square_meter="",
        total_price_per_square_meter="",
    )

    # There needs to be an empty row for sorting and filtering to work properly
    worksheet.append(empty_row)

    @cache
    def unwrap_range(cell_range: str) -> list[Any]:
        return [cell.value for row in worksheet[cell_range] for cell in row]

    @cache
    def conditional_range(value_range: str, **comparison_ranges_to_values: Any) -> list[Any]:
        """
        Returns values from `value_range` where the given comparison ranges
        contain all values as indicated by the mapping.
        """
        comparison_values: list[Any] = list(comparison_ranges_to_values.values())
        unwrapped_comparison_ranges = zip(*(unwrap_range(rang) for rang in comparison_ranges_to_values), strict=False)
        zipped_ranges = zip(unwrap_range(value_range), unwrapped_comparison_ranges, strict=True)
        return [
            value
            for value, range_values in zipped_ranges
            if all(range_value == comparison_values[i] for i, range_value in enumerate(range_values))
        ]

    summary_start = worksheet.max_row + 1

    summary_rows: list[SalesReportSummaryDefinition] = [
        SalesReportSummaryDefinition(
            title="Kaikki kaupat",
            subtitle="Lukumäärä",
            func=lambda x: len(unwrap_range(x)),
        ),
        SalesReportSummaryDefinition(
            subtitle="Keskiarvo",
            func=lambda x: mean(unwrap_range(x) or [0]),
        ),
        SalesReportSummaryDefinition(
            subtitle="Maksimi",
            func=lambda x: max(unwrap_range(x), default=0),
        ),
        SalesReportSummaryDefinition(
            subtitle="Minimi",
            func=lambda x: min(unwrap_range(x), default=0),
        ),
    ]

    for cost_area in range(1, 5):
        summary_rows += [
            None,  # empty row
            SalesReportSummaryDefinition(
                title=f"Kalleusalue {cost_area}",
                subtitle="Lukumäärä",
                func=lambda x, y=cost_area: len(conditional_range(x, **{f"A2:A{last_row}": y})),
            ),
            SalesReportSummaryDefinition(
                subtitle="Keskiarvo",
                func=lambda x, y=cost_area: mean(conditional_range(x, **{f"A2:A{last_row}": y}) or [0]),
            ),
            SalesReportSummaryDefinition(
                subtitle="Maksimi",
                func=lambda x, y=cost_area: max(conditional_range(x, **{f"A2:A{last_row}": y}), default=0),
            ),
            SalesReportSummaryDefinition(
                subtitle="Minimi",
                func=lambda x, y=cost_area: min(conditional_range(x, **{f"A2:A{last_row}": y}), default=0),
            ),
        ]

    sales_count_rows: list[int] = []

    for definition in summary_rows:
        if definition is None:
            worksheet.append(empty_row)
            continue

        if definition.subtitle == "Lukumäärä":
            sales_count_rows.append(worksheet.max_row + 1)

        worksheet.append(
            SalesReportColumns(
                cost_area="",
                postal_code="",
                apartment_address="",
                notification_date=definition.title,
                purchase_date=definition.subtitle,
                purchase_price_per_square_meter=definition.func(f"F2:F{last_row}"),
                total_price_per_square_meter=definition.func(f"G2:G{last_row}"),
            ),
        )

    euro_per_square_meter_format = "#,##0.00\\ \\€\\/\\m²"
    date_format = "DD.MM.YYYY"
    column_letters = string.ascii_uppercase[: len(column_headers)]

    format_sheet(
        worksheet,
        formatting_rules={
            # Add a border to the header row
            **{f"{letter}1": {"border": Border(bottom=Side(style="thin"))} for letter in column_letters},
            # Add a border to the last data row
            **{f"{letter}{last_row}": {"border": Border(bottom=Side(style="thin"))} for letter in column_letters},
            # Align the summary titles to the right
            **{
                f"E{summary_start + i}": {"alignment": Alignment(horizontal="right")}
                for i in range(0, len(summary_rows))
            },
            "B": {"alignment": Alignment(horizontal="right")},
            "D": {"number_format": date_format},
            "E": {"number_format": date_format},
            "F": {"number_format": euro_per_square_meter_format},
            "G": {"number_format": euro_per_square_meter_format},
            # Reset number format for sales count cells
            **{f"{letter}{row}": {"number_format": "General"} for row in sales_count_rows for letter in "FG"},
        },
    )

    resize_columns(worksheet)
    worksheet.protection.sheet = True
    return workbook


def build_sales_and_maximum_prices_report_excel(sales: list[ApartmentSale]) -> Workbook:
    workbook = Workbook()
    worksheet: Worksheet = workbook.active

    column_headers = SalesAndMaximumPricesReportColumns(
        cost_area="Kalleusalue",
        postal_code="Postinumero",
        apartment_address="Osoite",
        surface_area_square_meter="m2",
        purchase_date="Kauppapäivä",
        total_price="Toteutunut kauppahinta",
        total_price_per_square_meter="Kaupan neliöhinta",
        maximum_price="Enimmäishinta",
        maximum_price_per_square_meter="Enimmäishinnan neiöhinta",
    )
    worksheet.append(column_headers)

    # Prefetch surface area price ceilings
    surface_area_price_ceilings = {}
    for month_obj, value in SurfaceAreaPriceCeiling.objects.all().values_list("month", "value"):
        surface_area_price_ceilings[(month_obj.year, month_obj.month)] = value

    for sale in sales:
        if not sale.apartment.surface_area:
            raise ValidationError(
                detail={
                    api_settings.NON_FIELD_ERRORS_KEY: (
                        f"Surface area zero or missing for apartment {sale.apartment.address!r}. "
                        f"Cannot calculate price per square meter."
                    )
                },
            )

        # Pick calculation closest to the purchase date (latest calculation)
        # but only if it is valid on the purchase date.
        maximum_price_calculation = (
            sale.apartment.max_price_calculations.filter(
                calculation_date__lte=sale.purchase_date,
                valid_until__gte=sale.purchase_date,
            )
            .order_by("-calculation_date")
            .first()
        )

        maximum_price = None
        is_maximum_price_fallback = False
        if maximum_price_calculation is not None:
            maximum_price = maximum_price_calculation.maximum_price
        else:
            # Fall back to surface area price ceiling
            surface_area_price_ceiling = surface_area_price_ceilings.get(
                (sale.purchase_date.year, sale.purchase_date.month),
            )
            if surface_area_price_ceiling is not None:
                maximum_price = sale.apartment.surface_area * surface_area_price_ceiling
                is_maximum_price_fallback = True

        worksheet.append(
            SalesAndMaximumPricesReportColumns(
                cost_area=sale.apartment.postal_code.cost_area,
                postal_code=sale.apartment.postal_code.value,
                apartment_address=sale.apartment.address,
                surface_area_square_meter=sale.apartment.surface_area,
                purchase_date=sale.purchase_date,
                total_price=sale.total_price,
                total_price_per_square_meter=sale.total_price / sale.apartment.surface_area,
                maximum_price=maximum_price if maximum_price is not None else "",
                maximum_price_per_square_meter=(
                    (maximum_price / sale.apartment.surface_area) if maximum_price is not None else ""
                ),
            )
        )

        if is_maximum_price_fallback:
            maximum_price_cell = worksheet.cell(row=worksheet.max_row, column=8)
            maximum_price_per_square_meter_cell = worksheet.cell(row=worksheet.max_row, column=9)
            maximum_price_cell.font = Font(italic=True)
            maximum_price_per_square_meter_cell.font = Font(italic=True)

    last_row = worksheet.max_row
    worksheet.auto_filter.ref = worksheet.dimensions

    empty_row = SalesAndMaximumPricesReportColumns(
        cost_area="",
        postal_code="",
        apartment_address="",
        surface_area_square_meter="",
        purchase_date="",
        total_price="",
        total_price_per_square_meter="",
        maximum_price="",
        maximum_price_per_square_meter="",
    )

    # There needs to be an empty row for sorting and filtering to work properly
    worksheet.append(empty_row)

    @cache
    def unwrap_range(cell_range: str) -> list[Any]:
        return [cell.value for row in worksheet[cell_range] for cell in row]

    @cache
    def conditional_range(value_range: str, **comparison_ranges_to_values: Any) -> list[Any]:
        """
        Returns values from `value_range` where the given comparison ranges
        contain all values as indicated by the mapping.
        """
        comparison_values: list[Any] = list(comparison_ranges_to_values.values())
        unwrapped_comparison_ranges = zip(*(unwrap_range(rang) for rang in comparison_ranges_to_values), strict=False)
        zipped_ranges = zip(unwrap_range(value_range), unwrapped_comparison_ranges, strict=True)
        return [
            value
            for value, range_values in zipped_ranges
            if all(range_value == comparison_values[i] for i, range_value in enumerate(range_values))
        ]

    summary_start = worksheet.max_row + 1

    summary_rows: list[SalesReportSummaryDefinition] = [
        SalesReportSummaryDefinition(
            title="Kaikki kaupat",
            subtitle="Lukumäärä",
            func=lambda x: len(unwrap_range(x)),
        ),
        SalesReportSummaryDefinition(
            subtitle="Keskiarvo",
            func=lambda x: mean(unwrap_range(x) or [0]),
        ),
        SalesReportSummaryDefinition(
            subtitle="Maksimi",
            func=lambda x: max(unwrap_range(x), default=0),
        ),
        SalesReportSummaryDefinition(
            subtitle="Minimi",
            func=lambda x: min(unwrap_range(x), default=0),
        ),
    ]

    for cost_area in range(1, 5):
        summary_rows += [
            None,  # empty row
            SalesReportSummaryDefinition(
                title=f"Kalleusalue {cost_area}",
                subtitle="Lukumäärä",
                func=lambda x, y=cost_area: len(conditional_range(x, **{f"A2:A{last_row}": y})),
            ),
            SalesReportSummaryDefinition(
                subtitle="Keskiarvo",
                func=lambda x, y=cost_area: mean(conditional_range(x, **{f"A2:A{last_row}": y}) or [0]),
            ),
            SalesReportSummaryDefinition(
                subtitle="Maksimi",
                func=lambda x, y=cost_area: max(conditional_range(x, **{f"A2:A{last_row}": y}), default=0),
            ),
            SalesReportSummaryDefinition(
                subtitle="Minimi",
                func=lambda x, y=cost_area: min(conditional_range(x, **{f"A2:A{last_row}": y}), default=0),
            ),
        ]

    sales_count_rows: list[int] = []

    for definition in summary_rows:
        if definition is None:
            worksheet.append(empty_row)
            continue

        if definition.subtitle == "Lukumäärä":
            sales_count_rows.append(worksheet.max_row + 1)

        worksheet.append(
            SalesAndMaximumPricesReportColumns(
                cost_area="",
                postal_code="",
                apartment_address="",
                surface_area_square_meter=definition.title,
                purchase_date=definition.subtitle,
                total_price=definition.func(f"F2:F{last_row}"),
                total_price_per_square_meter=definition.func(f"G2:G{last_row}"),
                maximum_price="",
                maximum_price_per_square_meter="",
            ),
        )

    euro_format = "#,##0\\ \\€"
    euro_per_square_meter_format = "#,##0.00\\ \\€\\/\\m²"
    date_format = "DD.MM.YYYY"
    column_letters = string.ascii_uppercase[: len(column_headers)]

    format_sheet(
        worksheet,
        formatting_rules={
            # Add a border to the header row
            **{f"{letter}1": {"border": Border(bottom=Side(style="thin"))} for letter in column_letters},
            # Add a border to the last data row
            **{f"{letter}{last_row}": {"border": Border(bottom=Side(style="thin"))} for letter in column_letters},
            # Align the summary titles to the right
            **{
                f"E{summary_start + i}": {"alignment": Alignment(horizontal="right")}
                for i in range(0, len(summary_rows))
            },
            "B": {"alignment": Alignment(horizontal="right")},
            "E": {"number_format": date_format},
            "F": {"number_format": euro_format},
            "G": {"number_format": euro_per_square_meter_format},
            "H": {"number_format": euro_format},
            "I": {"number_format": euro_per_square_meter_format},
            # Reset number format for sales count cells
            **{f"{letter}{row}": {"number_format": "General"} for row in sales_count_rows for letter in "FG"},
        },
    )

    resize_columns(worksheet)
    worksheet.protection.sheet = True
    return workbook


def build_regulated_housing_companies_report_excel(
    housing_companies: list[HousingCompanyWithRegulatedReportAnnotations],
) -> Workbook:
    workbook = Workbook()
    worksheet: Worksheet = workbook.active

    column_headers = RegulatedHousingCompaniesReportColumns(
        cost_area="Kalleusalue",
        postal_code="Postinumero",
        housing_company_name="Yhtiö",
        address="Osoite",
        completion_date="Valmistumispäivä",
        apartment_count="Asuntojen lukumäärä",
        average_price_per_square_meter="Keskineliöhinta",
    )
    worksheet.append(column_headers)

    for housing_company in housing_companies:
        worksheet.append(
            RegulatedHousingCompaniesReportColumns(
                cost_area=housing_company.postal_code.cost_area,
                postal_code=housing_company.postal_code.value,
                housing_company_name=housing_company.display_name,
                address=housing_company.street_address,
                completion_date=housing_company.completion_date,
                apartment_count=housing_company.apartment_count,
                average_price_per_square_meter=housing_company.avg_price_per_square_meter,
            ),
        )

    last_row = worksheet.max_row
    worksheet.auto_filter.ref = worksheet.dimensions

    empty_row = RegulatedHousingCompaniesReportColumns(
        cost_area="",
        postal_code="",
        housing_company_name="",
        address="",
        completion_date="",
        apartment_count="",
        average_price_per_square_meter="",
    )

    # There needs to be an empty row for sorting and filtering to work properly
    worksheet.append(empty_row)

    euro_per_square_meter_format = "#,##0.00\\ \\€\\/\\m²"
    date_format = "DD.MM.YYYY"
    column_letters = string.ascii_uppercase[: len(column_headers)]

    format_sheet(
        worksheet,
        formatting_rules={
            # Add a border to the header row
            **{f"{letter}1": {"border": Border(bottom=Side(style="thin"))} for letter in column_letters},
            # Add a border to the last data row
            **{f"{letter}{last_row}": {"border": Border(bottom=Side(style="thin"))} for letter in column_letters},
            "B": {"alignment": Alignment(horizontal="right")},
            "E": {"number_format": date_format},
            "G": {"number_format": euro_per_square_meter_format},
        },
    )

    resize_columns(worksheet)
    worksheet.protection.sheet = True
    return workbook


def build_unregulated_housing_companies_report_excel(
    housing_companies: list[HousingCompanyWithUnregulatedReportAnnotations],
) -> Workbook:
    workbook = Workbook()
    worksheet: Worksheet = workbook.active

    column_headers = UnregulatedHousingCompaniesReportColumns(
        housing_company_name="Yhtiö",
        postal_code="Postinumero",
        completion_date="Valmistumispäivä",
        release_date="Vapautumispäivä",
        released_by_ploy_department="Vapautettu tontit yksikön toimesta?",
        apartment_count="Asuntojen lukumäärä",
    )
    worksheet.append(column_headers)

    for housing_company in housing_companies:
        worksheet.append(
            UnregulatedHousingCompaniesReportColumns(
                housing_company_name=housing_company.display_name,
                postal_code=housing_company.postal_code.value,
                completion_date=housing_company.completion_date,
                release_date=housing_company.release_date,
                released_by_ploy_department=(
                    "X" if housing_company.regulation_status == RegulationStatus.RELEASED_BY_PLOT_DEPARTMENT else ""
                ),
                apartment_count=housing_company.apartment_count,
            )
        )

    last_row = worksheet.max_row
    worksheet.auto_filter.ref = worksheet.dimensions

    empty_row = UnregulatedHousingCompaniesReportColumns(
        housing_company_name="",
        postal_code="",
        completion_date="",
        release_date="",
        released_by_ploy_department="",
        apartment_count="",
    )

    # There needs to be an empty row for sorting and filtering to work properly
    worksheet.append(empty_row)

    # Add summary rows
    worksheet.append(
        UnregulatedHousingCompaniesReportColumns(
            housing_company_name="Taloyhtiötä yhteensä",
            postal_code=len(housing_companies),
            completion_date="",
            release_date="",
            released_by_ploy_department="",
            apartment_count="",
        ),
    )
    worksheet.append(
        UnregulatedHousingCompaniesReportColumns(
            housing_company_name="Asuntoja yhteensä",
            postal_code=sum(housing_company.apartment_count for housing_company in housing_companies),
            completion_date="",
            release_date="",
            released_by_ploy_department="",
            apartment_count="",
        ),
    )

    date_format = "DD.MM.YYYY"
    column_letters = string.ascii_uppercase[: len(column_headers)]

    format_sheet(
        worksheet,
        formatting_rules={
            # Add a border to the header row
            **{f"{letter}1": {"border": Border(bottom=Side(style="thin"))} for letter in column_letters},
            # Add a border to the last data row
            **{f"{letter}{last_row}": {"border": Border(bottom=Side(style="thin"))} for letter in column_letters},
            "B": {"alignment": Alignment(horizontal="right")},
            "C": {"number_format": date_format},
            "D": {"number_format": date_format},
        },
    )

    resize_columns(worksheet)
    worksheet.protection.sheet = True
    return workbook


def build_property_managers_report_excel(
    housing_companies: list[HousingCompany],
    property_managers_with_no_housing_company: list[PropertyManager],
) -> Workbook:
    workbook = Workbook()
    worksheet: Worksheet = workbook.active

    column_headers = PropertyManagerReportColumns(
        property_manager_name="Isännöitsijätoimisto/isännöitsijä",
        property_manager_email="Sähköpostiosoite",
        housing_company_name="Taloyhtiön nimi",
    )
    worksheet.append(column_headers)

    for housing_company in housing_companies:
        worksheet.append(
            PropertyManagerReportColumns(
                property_manager_name=housing_company.property_manager.name,
                property_manager_email=housing_company.property_manager.email,
                housing_company_name=housing_company.display_name,
            )
        )

    for property_manager in property_managers_with_no_housing_company:
        worksheet.append(
            PropertyManagerReportColumns(
                property_manager_name=property_manager.name,
                property_manager_email=property_manager.email,
                housing_company_name="",
            )
        )

    last_row = worksheet.max_row
    worksheet.auto_filter.ref = worksheet.dimensions

    column_letters = string.ascii_uppercase[: len(column_headers)]

    format_sheet(
        worksheet,
        formatting_rules={
            # Add a border to the header row
            **{f"{letter}1": {"border": Border(bottom=Side(style="thin"))} for letter in column_letters},
            # Add a border to the last data row
            **{f"{letter}{last_row}": {"border": Border(bottom=Side(style="thin"))} for letter in column_letters},
        },
    )

    resize_columns(worksheet)
    worksheet.protection.sheet = True
    return workbook


def build_housing_company_state_report_excel(
    housing_companies: list[HousingCompanyWithStateReportAnnotations],
) -> Workbook:
    workbook = Workbook()
    worksheet: Worksheet = workbook.active

    column_headers = HousingCompanyStatesReportColumns(
        state="Taloyhtiön tila",
        housing_company_count="Taloyhtiöiden lukumäärä",
        apartment_count="Asuntojen lukumäärä",
    )
    worksheet.append(column_headers)

    states = sort_housing_companies_by_state(housing_companies)

    for state, counts in states.items():
        worksheet.append(
            HousingCompanyStatesReportColumns(
                state=state.value,
                housing_company_count=counts["housing_company_count"],
                apartment_count=counts["apartment_count"],
            )
        )

    last_row = worksheet.max_row
    worksheet.auto_filter.ref = worksheet.dimensions

    empty_row = HousingCompanyStatesReportColumns(
        state="",
        housing_company_count="",
        apartment_count="",
    )

    # There needs to be an empty row for sorting and filtering to work properly
    worksheet.append(empty_row)

    # Add summary rows
    worksheet.append(
        HousingCompanyStatesReportColumns(
            state="Yhtiöitä yhteensä",
            housing_company_count=len(housing_companies),
            apartment_count="",
        ),
    )
    worksheet.append(
        HousingCompanyStatesReportColumns(
            state="Asuntoja yhteensä",
            housing_company_count=sum(housing_company.apartment_count for housing_company in housing_companies),
            apartment_count="",
        ),
    )

    column_letters = string.ascii_uppercase[: len(column_headers)]

    format_sheet(
        worksheet,
        formatting_rules={
            # Add a border to the header row
            **{f"{letter}1": {"border": Border(bottom=Side(style="thin"))} for letter in column_letters},
            # Add a border to the last data row
            **{f"{letter}{last_row}": {"border": Border(bottom=Side(style="thin"))} for letter in column_letters},
        },
    )

    resize_columns(worksheet)
    worksheet.protection.sheet = True
    return workbook


def sort_housing_companies_by_state(
    housing_companies: list[HousingCompanyWithStateReportAnnotations],
) -> dict[ReportState, HousingCompanyStatesCount]:
    states: dict[ReportState, HousingCompanyStatesCount] = {
        name: HousingCompanyStatesCount(
            housing_company_count=0,
            apartment_count=0,
        )
        for name in ReportState
    }

    for housing_company in housing_companies:
        if housing_company.hitas_type == HitasType.HALF_HITAS:
            states[ReportState.HALF_HITAS]["housing_company_count"] += 1
            states[ReportState.HALF_HITAS]["apartment_count"] += housing_company.apartment_count

        elif housing_company.completion_date is None:
            states[ReportState.NOT_READY]["housing_company_count"] += 1
            states[ReportState.NOT_READY]["apartment_count"] += housing_company.apartment_count

        elif housing_company.regulation_status == RegulationStatus.REGULATED:
            states[ReportState.REGULATED]["housing_company_count"] += 1
            states[ReportState.REGULATED]["apartment_count"] += housing_company.apartment_count

        elif housing_company.regulation_status == RegulationStatus.RELEASED_BY_HITAS:
            states[ReportState.RELEASED_BY_HITAS]["housing_company_count"] += 1
            states[ReportState.RELEASED_BY_HITAS]["apartment_count"] += housing_company.apartment_count

        elif housing_company.regulation_status == RegulationStatus.RELEASED_BY_PLOT_DEPARTMENT:
            states[ReportState.RELEASED_BY_PLOT_DEPARTMENT]["housing_company_count"] += 1
            states[ReportState.RELEASED_BY_PLOT_DEPARTMENT]["apartment_count"] += housing_company.apartment_count

    return states


def build_sales_by_postal_code_and_area_report_excel(sales: list[ApartmentSale]) -> Workbook:
    workbook = Workbook()
    worksheet: Worksheet = workbook.active

    column_headers = SalesByCostAreaColumns(
        cost_area="Kalleusalue",
        postal_code="Postinumero",
        room_label="Huoneluku",
        sales_count="Kauppojen lukumäärä",
        average_price_per_square_meter="Keskineliöhinta",
        minimum_price_per_square_meter="Alin neliöhinta",
        maximum_price_per_square_meter="Ylin neliöhinta",
    )
    worksheet.append(column_headers)

    results = sort_sales_by_cost_area(sales)

    for cost_area, sales_info_by_room_label_by_postal_code in results.sales_by_cost_area.items():
        for postal_code, sales_info_by_room_label in sales_info_by_room_label_by_postal_code.items():
            for room_label, sales_info in sales_info_by_room_label.items():
                worksheet.append(
                    SalesByCostAreaColumns(
                        cost_area=cost_area,
                        postal_code=postal_code,
                        room_label=room_label,
                        sales_count=sales_info["sales_count"],
                        average_price_per_square_meter=sales_info["sum"] / sales_info["sales_count"],
                        minimum_price_per_square_meter=sales_info["minimum"],
                        maximum_price_per_square_meter=sales_info["maximum"],
                    )
                )

    last_row = worksheet.max_row
    worksheet.auto_filter.ref = worksheet.dimensions

    empty_row = SalesByCostAreaColumns(
        cost_area="",
        postal_code="",
        room_label="",
        sales_count="",
        average_price_per_square_meter="",
        minimum_price_per_square_meter="",
        maximum_price_per_square_meter="",
    )

    # There needs to be an empty row for sorting and filtering to work properly
    worksheet.append(empty_row)

    # Add summary rows
    worksheet.append(
        SalesByCostAreaColumns(
            cost_area="",
            postal_code="",
            room_label="Koko Helsingin alue",
            sales_count=results.overall_count,
            average_price_per_square_meter=results.overall_average,
            minimum_price_per_square_meter=results.overall_minimum,
            maximum_price_per_square_meter=results.overall_maximum,
        )
    )

    euro_per_square_meter_format = "#,##0.00\\ \\€\\/\\m²"
    column_letters = string.ascii_uppercase[: len(column_headers)]

    format_sheet(
        worksheet,
        formatting_rules={
            # Add a border to the header row
            **{f"{letter}1": {"border": Border(bottom=Side(style="thin"))} for letter in column_letters},
            # Add a border to the last data row
            **{f"{letter}{last_row}": {"border": Border(bottom=Side(style="thin"))} for letter in column_letters},
            "B": {"alignment": Alignment(horizontal="right")},
            "C": {"alignment": Alignment(horizontal="right")},
            "E": {"number_format": euro_per_square_meter_format},
            "F": {"number_format": euro_per_square_meter_format},
            "G": {"number_format": euro_per_square_meter_format},
        },
    )

    resize_columns(worksheet)
    worksheet.protection.sheet = True
    return workbook


def sort_sales_by_cost_area(sales: list[ApartmentSale]) -> SalesByCostArea:
    sales_by_cost_area: dict[CostAreaT, dict[PostalCodeT, SalesInfoByRoomCount]] = {}
    rooms_to_label = {1: "1h", 2: "2h"}  # >=3 not found -> sorted to 3h+
    overall_count = 0
    overall_average = Decimal("0")
    overall_minimum = Decimal("inf")
    overall_maximum = Decimal("-inf")
    for sale in sales:
        if not sale.apartment.surface_area:
            raise ValidationError(
                detail={
                    api_settings.NON_FIELD_ERRORS_KEY: (
                        f"Surface area zero or missing for apartment {sale.apartment.address!r}. "
                        f"Cannot calculate price per square meter."
                    )
                },
            )

        cost_area: CostAreaT = sale.apartment.postal_code.cost_area  # type: ignore
        postal_code: PostalCodeT = sale.apartment.postal_code.value
        price_per_square_meter: Decimal = sale.total_price / sale.apartment.surface_area
        room_label: RoomLabelT = rooms_to_label.get(sale.apartment.rooms, "3h+")

        sales_by_cost_area.setdefault(cost_area, {})
        sales_by_cost_area[cost_area].setdefault(postal_code, {})
        sales_by_cost_area[cost_area][postal_code].setdefault(
            room_label,
            SalesInfo(
                sales_count=0,
                sum=Decimal("0"),
                minimum=Decimal("inf"),
                maximum=Decimal("-inf"),
            ),
        )

        sales_by_cost_area[cost_area][postal_code][room_label]["sales_count"] += 1
        sales_by_cost_area[cost_area][postal_code][room_label]["sum"] += price_per_square_meter
        if price_per_square_meter < sales_by_cost_area[cost_area][postal_code][room_label]["minimum"]:
            sales_by_cost_area[cost_area][postal_code][room_label]["minimum"] = price_per_square_meter
        if price_per_square_meter > sales_by_cost_area[cost_area][postal_code][room_label]["maximum"]:
            sales_by_cost_area[cost_area][postal_code][room_label]["maximum"] = price_per_square_meter

        overall_count += 1
        overall_average += price_per_square_meter
        if price_per_square_meter < overall_minimum:
            overall_minimum = price_per_square_meter
        if price_per_square_meter > overall_maximum:
            overall_maximum = price_per_square_meter

    overall_average /= overall_count

    return SalesByCostArea(
        sales_by_cost_area=sales_by_cost_area,
        overall_count=overall_count,
        overall_average=overall_average,
        overall_minimum=overall_minimum,
        overall_maximum=overall_maximum,
    )


def build_regulated_ownerships_report_excel(ownerships: list[Ownership]) -> Workbook:
    workbook = Workbook()
    worksheet: Worksheet = workbook.active

    column_headers = OwnershipReportColumns(
        owner_name="Omistajan nimi",
        apartment_address="Asunnon osoite",
        postal_code="Postinumero",
        owner_identifier="Omistajan henkilö- tai Y-tunnus",
        housing_company_name="Yhtiön nimi",
        housing_company_completion_date="Yhtiön valmistumispäivä",
        cost_area="Kalleusalue",
    )
    worksheet.append(column_headers)

    # Cache completion_date as it is the most expensive operation here
    # because it queries all apartment completion dates on each iteration
    # and there are relatively few housing companies.
    completion_dates_by_housing_company_id = {}

    for ownership in ownerships:
        try:
            completion_date = completion_dates_by_housing_company_id[
                ownership.apartment.building.real_estate.housing_company.pk
            ]
        except KeyError:
            completion_date = ownership.apartment.building.real_estate.housing_company.completion_date
            completion_dates_by_housing_company_id[ownership.apartment.building.real_estate.housing_company.pk] = (
                completion_date
            )
        worksheet.append(
            OwnershipReportColumns(
                owner_name=Owner.OBFUSCATED_OWNER_NAME if ownership.owner.non_disclosure else ownership.owner.name,
                apartment_address=ownership.apartment.address,
                postal_code=ownership.apartment.postal_code.value,
                owner_identifier="" if ownership.owner.non_disclosure else ownership.owner.identifier,
                housing_company_name=ownership.apartment.building.real_estate.housing_company.display_name,
                housing_company_completion_date=completion_date,
                cost_area=ownership.apartment.postal_code.cost_area,
            )
        )

    _basic_format_sheet(column_headers, worksheet)
    return workbook


def build_multiple_ownerships_report_excel(ownerships: list[OwnershipWithApartmentCount]) -> Workbook:
    workbook = Workbook()
    worksheet: Worksheet = workbook.active

    column_headers = MultipleOwnershipReportColumns(
        owner_name="Omistajan nimi",
        apartment_address="Asunnon osoite",
        postal_code="Postinumero",
        apartment_count="Omistajan asuntojen lukumäärä",
        owner_identifier="Omistajan henkilö- tai Y-tunnus",
        housing_company_name="Yhtiön nimi",
        housing_company_completion_date="Yhtiön valmistumispäivä",
        cost_area="Kalleusalue",
    )
    worksheet.append(column_headers)

    for ownership in ownerships:
        worksheet.append(
            MultipleOwnershipReportColumns(
                owner_name=Owner.OBFUSCATED_OWNER_NAME if ownership.owner.non_disclosure else ownership.owner.name,
                apartment_address=ownership.apartment.address,
                postal_code=ownership.apartment.postal_code.value,
                apartment_count=ownership.apartment_count,
                owner_identifier="" if ownership.owner.non_disclosure else ownership.owner.identifier,
                housing_company_name=ownership.apartment.building.real_estate.housing_company.display_name,
                housing_company_completion_date=ownership.apartment.building.real_estate.housing_company.completion_date,
                cost_area=ownership.apartment.postal_code.cost_area,
            )
        )

    _basic_format_sheet(column_headers, worksheet)
    return workbook


def build_owners_by_housing_companies_report_excel(entries) -> Workbook:
    workbook = Workbook()
    worksheet: Worksheet = workbook.active
    column_headers = OwnersByHousingCompanyReportColumns(
        number="Asunnon nro",
        surface_area="Asunnon pinta-ala",
        share_numbers="Osakenumerot",
        purchase_date="Kauppakirjapäivä",
        owner_name="Omistajan nimi",
        owner_ssn="Henkilötunnus",
    )

    worksheet.append(column_headers)
    for entry in entries:
        worksheet.append(
            OwnersByHousingCompanyReportColumns(
                number=entry.sale.apartment.apartment_number,
                surface_area=entry.sale.apartment.surface_area,
                share_numbers=f"{entry.sale.apartment.share_number_start}-{entry.sale.apartment.share_number_end}",
                purchase_date=entry.sale.purchase_date,
                owner_name=Owner.OBFUSCATED_OWNER_NAME if entry.owner.non_disclosure else entry.owner.name,
                owner_ssn=" " if entry.owner.non_disclosure else entry.owner.identifier,
            )
        )

    _basic_format_sheet(column_headers, worksheet)
    return workbook


def _basic_format_sheet(
    column_headers: Union[OwnersByHousingCompanyReportColumns, MultipleOwnershipReportColumns], worksheet: Worksheet
) -> None:
    column_letters = string.ascii_uppercase[: len(column_headers)]
    last_row = worksheet.max_row
    worksheet.auto_filter.ref = worksheet.dimensions

    format_sheet(
        worksheet,
        formatting_rules={
            # Add a border to the header row
            **{f"{letter}1": {"border": Border(bottom=Side(style="thin"))} for letter in column_letters},
            # Add a border to the last data row
            **{f"{letter}{last_row}": {"border": Border(bottom=Side(style="thin"))} for letter in column_letters},
        },
    )

    resize_columns(worksheet)
    worksheet.protection.sheet = True
