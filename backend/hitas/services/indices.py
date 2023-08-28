import datetime
import logging
from decimal import Decimal
from typing import Literal, NamedTuple, Optional, Union

from dateutil.relativedelta import relativedelta
from django.db.models import Case, F, OuterRef, Q, Subquery, Value, When
from django.db.models.functions import Coalesce, NullIf, Round
from django.utils import timezone
from openpyxl.styles import Alignment, Border, Side
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from hitas.calculations.depreciation_percentage import depreciation_multiplier
from hitas.calculations.helpers import months_between_dates
from hitas.exceptions import HitasModelNotFound, ModelConflict
from hitas.models import (
    ConstructionPriceIndex,
    ConstructionPriceIndex2005Equal100,
    SurfaceAreaPriceCeiling,
)
from hitas.models._base import HitasModelDecimalField
from hitas.models.housing_company import HitasType, HousingCompanyWithAnnotations
from hitas.models.indices import (
    CalculationData,
    HousingCompanyData,
    MarketPriceIndex,
    MarketPriceIndex2005Equal100,
    SurfaceAreaPriceCeilingCalculationData,
    SurfaceAreaPriceCeilingResult,
)
from hitas.services.housing_company import get_completed_housing_companies, make_index_adjustment_for_housing_companies
from hitas.utils import format_sheet, hitas_calculation_quarter, monthify, resize_columns, roundup

logger = logging.getLogger()


class ReportColumns(NamedTuple):
    display_name: str
    acquisition_price: Decimal | str
    indices: str
    change: Decimal | str
    adjusted_acquisition_price: Decimal | str
    surface_area: Decimal | str
    price_per_square_meter: Decimal | str


def calculate_surface_area_price_ceiling(calculation_date: datetime.date) -> list[SurfaceAreaPriceCeilingResult]:
    calculation_month = hitas_calculation_quarter(calculation_date)

    logger.info(f"Calculating surface area price ceiling for {calculation_month.isoformat()!r}...")

    logger.info("Fetching housing companies...")
    housing_companies = get_completed_housing_companies(
        completion_month=calculation_month,
        include_excluded_from_statistics=False,
        include_rental_hitas=False,
        include_half_hitas=False,
    )

    if not housing_companies:
        raise ModelConflict(
            f"No regulated housing companies completed before {calculation_month.isoformat()!r}.",
            error_code="missing",
        )

    unadjusted_prices: dict[str, Optional[Decimal]] = {
        housing_company.uuid.hex: housing_company.avg_price_per_square_meter for housing_company in housing_companies
    }

    logger.info("Making index adjustments...")
    indices = make_index_adjustment_for_housing_companies(housing_companies, calculation_month)

    logger.info(
        f"Setting surface area price ceiling for the next three months "
        f"starting from {calculation_month.isoformat()!r}..."
    )
    total = _calculate_surface_area_price_ceiling_for_housing_companies(housing_companies)
    results = _create_surface_area_price_ceiling_for_next_three_months(calculation_month, total)

    logger.info("Saving calculation data for later use...")
    _save_calculation_data(calculation_month, housing_companies, indices, unadjusted_prices, results)

    logger.info("Surface area price ceiling calculation complete!")
    return results


def _calculate_surface_area_price_ceiling_for_housing_companies(
    housing_companies: list[HousingCompanyWithAnnotations],
) -> Decimal:
    total: Decimal = Decimal(sum((housing_company.avg_price_per_square_meter for housing_company in housing_companies)))
    return roundup(total / len(housing_companies), precision=0)


def _create_surface_area_price_ceiling_for_next_three_months(
    from_: datetime.date,
    value: Decimal,
) -> list[SurfaceAreaPriceCeilingResult]:
    results: list[SurfaceAreaPriceCeilingResult] = []
    for month in range(3):
        surface_area_price_ceiling, _ = SurfaceAreaPriceCeiling.objects.update_or_create(
            month=from_ + relativedelta(months=month),
            defaults={"value": value},
        )
        results.append(
            SurfaceAreaPriceCeilingResult(
                month=surface_area_price_ceiling.month.isoformat()[:-3],
                value=float(surface_area_price_ceiling.value),
            )
        )
    return results


def _save_calculation_data(
    calculation_month: datetime.date,
    housing_companies: list[HousingCompanyWithAnnotations],
    indices: dict[Literal["old", "new"], dict[datetime.date, Decimal]],
    unadjusted_prices: dict[str, Optional[Decimal]],
    surface_area_price_ceilings: list[SurfaceAreaPriceCeilingResult],
) -> tuple[SurfaceAreaPriceCeilingCalculationData, bool]:
    data = CalculationData(
        housing_company_data=[
            HousingCompanyData(
                name=housing_company.display_name,
                completion_date=housing_company.completion_date.isoformat(),
                surface_area=float(housing_company.surface_area),
                realized_acquisition_price=float(housing_company.realized_acquisition_price),
                unadjusted_average_price_per_square_meter=float(unadjusted_prices[housing_company.uuid.hex]),
                adjusted_average_price_per_square_meter=float(housing_company.avg_price_per_square_meter),
                completion_month_index=float(indices[key][housing_company.completion_month]),
                calculation_month_index=float(indices[key][calculation_month]),
            )
            for housing_company in housing_companies
            if (key := "old" if housing_company.hitas_type.old_hitas_ruleset else "new")  # NOSONAR
        ],
        created_surface_area_price_ceilings=surface_area_price_ceilings,
    )
    return SurfaceAreaPriceCeilingCalculationData.objects.update_or_create(
        calculation_month=calculation_month,
        defaults={"data": data},
    )


def subquery_appropriate_cpi(outer_ref: str, housing_company_ref: str) -> Case:
    """
    Make a subquery for appropriate construction price index based on housing company's financing method.
    """
    return Case(
        When(
            condition=~Q(**{f"{housing_company_ref}__hitas_type__in": HitasType.with_new_hitas_ruleset()}),
            then=(
                Subquery(
                    queryset=(ConstructionPriceIndex.objects.filter(month=OuterRef(outer_ref)).values("value")[:1]),
                    output_field=HitasModelDecimalField(),
                )
            ),
        ),
        default=(
            Subquery(
                queryset=(
                    ConstructionPriceIndex2005Equal100.objects.filter(month=OuterRef(outer_ref)).values("value")[:1]
                ),
                output_field=HitasModelDecimalField(),
            )
        ),
    )


def get_surface_area_price_ceiling_results(calculation_date: datetime.date) -> SurfaceAreaPriceCeilingCalculationData:
    try:
        return SurfaceAreaPriceCeilingCalculationData.objects.get(
            calculation_month=hitas_calculation_quarter(calculation_date),
        )
    except SurfaceAreaPriceCeilingCalculationData.DoesNotExist as error:
        raise HitasModelNotFound(SurfaceAreaPriceCeilingCalculationData) from error


def build_surface_area_price_ceiling_report_excel(results: SurfaceAreaPriceCeilingCalculationData) -> Workbook:
    workbook = Workbook()
    worksheet: Worksheet = workbook.active

    columns = ReportColumns(
        display_name="Yhtiö",
        acquisition_price="Hankinta-arvo",
        indices="Indeksit",
        change="Muutos",
        adjusted_acquisition_price="Takistettu hinta",
        surface_area="Pinta-ala",
        price_per_square_meter="E-hinta/m²",
    )
    worksheet.append(columns)

    for housing_company in results.data["housing_company_data"]:
        unadjusted = Decimal(housing_company["unadjusted_average_price_per_square_meter"])
        adjusted = Decimal(housing_company["adjusted_average_price_per_square_meter"])
        surface_area = Decimal(housing_company["surface_area"])
        data = ReportColumns(
            display_name=housing_company["name"],
            acquisition_price=Decimal(housing_company["realized_acquisition_price"]),
            indices=f"{housing_company['completion_month_index']}/{housing_company['calculation_month_index']}",
            change=(adjusted * surface_area) - (unadjusted * surface_area),
            adjusted_acquisition_price=adjusted * surface_area,
            surface_area=surface_area,
            price_per_square_meter=adjusted,
        )
        worksheet.append(data)

    last_row = worksheet.max_row
    worksheet.auto_filter.ref = worksheet.dimensions

    # There needs to be an empty row for sorting and filtering to work properly
    worksheet.append(
        ReportColumns(
            display_name="",
            acquisition_price="",
            indices="",
            change="",
            adjusted_acquisition_price="",
            surface_area="",
            price_per_square_meter="",
        )
    )

    summary_start = worksheet.max_row + 1
    summary_rows = {"Summa": "SUM", "Keskiarvo": "AVERAGE"}
    for title, formula in summary_rows.items():
        worksheet.append(
            ReportColumns(
                display_name=title,
                acquisition_price=f"={formula}(B2:B{last_row})",
                indices="",
                change=f"={formula}(D2:D{last_row})",
                adjusted_acquisition_price=f"={formula}(E2:E{last_row})",
                surface_area=f"={formula}(F2:F{last_row})",
                price_per_square_meter=f"={formula}(G2:G{last_row})",
            )
        )
    summary_end = worksheet.max_row

    worksheet.append(
        ReportColumns(
            display_name="Rajaneliöhinta",
            acquisition_price=Decimal(results.data["created_surface_area_price_ceilings"][0]["value"]),
            indices="",
            change="",
            adjusted_acquisition_price="",
            surface_area="",
            price_per_square_meter="",
        )
    )

    euro_format = "#,##0.00\\ €"
    square_meter_format = "#,##0.00\\ \\m\\²"

    format_sheet(
        worksheet,
        formatting_rules={
            # Add a border to the header row
            **{f"{letter}1": {"border": Border(bottom=Side(style="thin"))} for letter in "ABCDEFG"},
            # Add a border to the last data row
            **{f"{letter}{last_row}": {"border": Border(bottom=Side(style="thin"))} for letter in "ABCDEFG"},
            # Align the summary titles to the right
            **{
                f"A{summary_start + i}": {"alignment": Alignment(horizontal="right")}
                for i in range(0, len(summary_rows) + 1)  # additional +1 for "Rajaneliöhinta" row
            },
            # Add border to the end of the summary row
            **{f"{letter}{summary_end}": {"border": Border(bottom=Side(style="thin"))} for letter in "BCDEFG"},
            # Last summary row needs both underline and alignment
            f"A{summary_end}": {
                "border": Border(bottom=Side(style="thin")),
                "alignment": Alignment(horizontal="right"),
            },
            "B": {"number_format": euro_format},
            "C": {"alignment": Alignment(horizontal="right")},
            "D": {"number_format": euro_format},
            "E": {"number_format": euro_format},
            "F": {"number_format": square_meter_format},
            "G": {"number_format": euro_format},
        },
    )

    resize_columns(worksheet)
    worksheet.protection.sheet = True
    return workbook


def subquery_apartment_current_surface_area_price(
    calculation_date: Optional[datetime.date] = None,
) -> Round:
    calculation_date = timezone.now().date() if calculation_date is None else calculation_date
    calculation_month = monthify(calculation_date)

    current_value = Subquery(
        SurfaceAreaPriceCeiling.objects.filter(month=calculation_month).values("value"),
        output_field=HitasModelDecimalField(),
    )

    return Round(
        F("surface_area") * current_value,
        precision=2,
    )


def subquery_apartment_first_sale_acquisition_price_index_adjusted(
    table: Union[
        type[MarketPriceIndex],
        type[MarketPriceIndex2005Equal100],
        type[ConstructionPriceIndex],
        type[ConstructionPriceIndex2005Equal100],
    ],
    completion_date: Optional[datetime.date],
    calculation_date: datetime.date,
) -> Union[Round | Value]:
    """
    If 'completion_date' is missing, calculating index for that month will fail
    and index price will be null, so we can skip this calculation freely

    Requires `completion_month` to be annotated to the queryset
    """
    if completion_date is None:
        return Value(None, output_field=HitasModelDecimalField())

    calculation_date = timezone.now().date() if calculation_date is None else calculation_date
    calculation_month = monthify(calculation_date)

    original_value = Subquery(
        table.objects.filter(month=OuterRef("completion_month")).values("value"),
        output_field=HitasModelDecimalField(),
    )
    current_value = Subquery(
        table.objects.filter(month=calculation_month).values("value"),
        output_field=HitasModelDecimalField(),
    )

    # Initialize default values
    depreciation: Value = Value(1, output_field=HitasModelDecimalField())
    interest: Union[Case, Value, Coalesce] = Value(0, output_field=HitasModelDecimalField())

    # Override values based on table type for Old Hitas ruleset
    if issubclass(table, MarketPriceIndex):
        interest = Coalesce(F("interest_during_construction_mpi"), 0, output_field=HitasModelDecimalField())
    elif issubclass(table, ConstructionPriceIndex):
        depreciation = Value(
            depreciation_multiplier(months_between_dates(completion_date, calculation_date)),
            output_field=HitasModelDecimalField(),
        )
        interest = Case(
            When(
                condition=Q(completion_date__lt=datetime.date(2005, 1, 1)),
                then=Coalesce(F("interest_during_construction_cpi"), 0, output_field=HitasModelDecimalField()),
            ),
            default=Coalesce(F("interest_during_construction_mpi"), 0, output_field=HitasModelDecimalField()),
            output_field=HitasModelDecimalField(),
        )

    return Round(
        (
            F("_first_sale_purchase_price")
            + F("_first_sale_share_of_housing_company_loans")
            + F("additional_work_during_construction")
            + interest
        )
        * depreciation
        * current_value
        / NullIf(original_value, 0, output_field=HitasModelDecimalField()),  # prevent zero division errors
        precision=2,
    )
