from io import BytesIO
from tempfile import NamedTemporaryFile
from typing import Any, Literal, Optional, Protocol, TypeAlias

from django.core.handlers.wsgi import WSGIRequest
from django.http import HttpResponse
from openpyxl.cell import Cell
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from rest_framework.exceptions import ErrorDetail, ValidationError
from rest_framework.parsers import BaseParser
from rest_framework.serializers import Serializer


def get_excel_response(filename: str, excel: Workbook) -> HttpResponse:
    with NamedTemporaryFile() as tmp:
        excel.save(tmp.name)
        tmp.seek(0)
        data: bytes = tmp.read()

    response = HttpResponse(data, content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    return response


class _ExcelParser(BaseParser):
    def parse(
        self,
        stream: WSGIRequest,
        media_type: Optional[str] = None,
        parser_context: Optional[dict[str, Any]] = None,
    ) -> Workbook:
        return parse_excel_from_bytes(request=stream)


class OldExcelParser(_ExcelParser):
    media_type = "application/vnd.ms-excel"  # .xls


class NewExcelParser(_ExcelParser):
    media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"  # .xlsx


def parse_excel_from_bytes(request: WSGIRequest) -> Workbook:
    return load_workbook(filename=BytesIO(request.read()), data_only=True)


# fmt: off
# These aren't the only valid values but will provide valuable typing hints in most cases
RowNumber: TypeAlias = Literal["1", "2", "3", "4", "5", "6", "7", "8", "9", "final"]
ColumnLetter: TypeAlias = Literal[
    "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N",
    "O", "P", "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z"
]
# fmt: on

FieldName: TypeAlias = str
RowFormat: TypeAlias = dict[FieldName, ColumnLetter]
ExtraFormat: TypeAlias = dict[RowNumber, dict[FieldName, ColumnLetter]]
ErrorText: TypeAlias = str
ErrorData: TypeAlias = dict[FieldName, list[ErrorText]]


class RowPostValidator(Protocol):
    def __call__(self, row_data: list[dict[str, Any]], row_format: RowFormat, errors: ErrorData) -> None:
        """Validator function"""


def parse_sheet(  # NOSONAR
    worksheet: Worksheet,
    *,
    row_data_key: str,
    row_format: RowFormat,
    extra_format: ExtraFormat,
    row_serializer: type[Serializer],
    extra_serializer: type[Serializer],
    row_post_validators: list[RowPostValidator],
) -> dict[str, Any]:
    """Parse an Excel sheet using json instructions.

    :param worksheet: Worksheet to parse.
    :param row_format: Mapping of field name to column letter for row data.
    :param row_serializer: Serializer for data rows.
    :param extra_format: Mapping of row numbers (1-indexed) to dicts of field names to column letters.
                         These fields will be processed separately from the row data.
                         The processing will move on to row data once a row number cannot be found in this mapping.
                         Use a special "final" key to list columns which will have data on the final row
                         (e.g. column sums) so that the final row can be identified properly, and the extra data
                         can be included for 'extra_serializer'.
    :param extra_serializer: Serializer for extra data found from the sheet using 'extra_format'.
                             When validating, the row data can be found from 'self.initial_data' under 'row_data_key'.
    :param row_data_key: Key to return the row data under in the final response.
    :param row_post_validators: Extra validators to run for all rows after they have been parsed.
                                Errors should be added to the 'error'-dictionary passed to the validator.
                                'row_serializer' should contain a field for "row" so that post validators
                                can use it to format their errors with the correct cell indicator.
    """

    non_empty_columns: set[int] = {
        column_index_from_string(letter) - 1 for letter in extra_format.get("final", {}).values()
    }
    field_name_to_cell: dict[str, str] = {
        field_name: f"{column}{row_number}" if row_number != "final" else f"{column}"
        for row_number, rules in extra_format.items()
        for field_name, column in rules.items()
    }

    errors: dict[str, list[str]] = {}
    row_data: list[dict[str, Any]] = []
    extra_data: dict[str, Any] = {}

    cells: tuple[Cell, ...]
    for cells in worksheet.iter_rows():
        row_number: RowNumber = str(cells[0].row)  # type: ignore

        rules = extra_format.get(row_number)
        if rules is not None:
            for key, column in rules.items():
                extra_data[key] = cells[column_index_from_string(column) - 1].value
            continue

        final_row = all(cells[i].value is None for i in range(len(cells)) if i not in non_empty_columns)
        if final_row:
            for key, column in extra_format.get("final", {}).items():
                field_name_to_cell[key] += row_number  # add row number for error formatting
                extra_data[key] = cells[column_index_from_string(column) - 1].value
            break

        row_data.append(_parse_row(cells, row_format, row_serializer, errors))

    # Row post validation
    for validator in row_post_validators:
        validator(row_data, row_format, errors)

    # Add row data so that extra serializer can access it if needed
    extra_data[row_data_key] = row_data

    # Extra data validation
    extra = extra_serializer(data=extra_data)
    if not extra.is_valid():
        field_name: str
        error: list[ErrorDetail]
        for field_name, error in extra.errors.items():
            errors[f"{field_name_to_cell[field_name]}.{field_name}"] = error

    extra_data = extra.data
    extra_data[row_data_key] = row_data

    if errors:
        raise ValidationError(errors)

    return extra_data


def _parse_row(
    cells: tuple[Cell, ...],
    row_format: RowFormat,
    row_serializer: type[Serializer],
    errors: ErrorData,
) -> dict[str, Any]:
    row_data: dict[str, Any] = {
        field_name: cells[column_index_from_string(column) - 1].value for field_name, column in row_format.items()
    }
    row_data["row"] = row_number = cells[0].row  # add row number for post validators

    row = row_serializer(data=row_data)
    if row.is_valid():
        return row.data

    field_name: str
    error: list[ErrorDetail]
    for field_name, error in row.errors.items():
        errors[error_key(field_name, row_number, row_format)] = error

    return row_data  # 'row.data' might not be valid


def error_key(field_name: FieldName, row: int, row_format: RowFormat) -> str:
    return f"{row_format[field_name]}{row}.{field_name}"
